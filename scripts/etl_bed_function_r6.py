#!/usr/bin/env python3
"""
phase5 ETL: 病床機能報告 令和6年度 (R6=2024) → medical_areas_national.json 更新

R1 (2019) → R6 (2024) で5年分の更新。
データ出典: 厚労省 令和6年度病床機能報告公表データ (2025-09-30公開)
URL: https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/open_data_00018.html

R6スキーマ変更:
  R1: 1ファイル (病棟票+施設票の2シート)
  R6: 施設票×1 + 診療所票×1 + 様式1_病棟票×7地域 + 様式2×7地域

本ETLは「様式1_病棟票」7ファイルから医療圏単位で集計:
  hosp  = 医療圏内のユニーク医療機関コード数
  wards = 医療圏内の病棟行数
  beds  = 医療圏内の (一般病床_許可病床 + 療養病床_許可病床) 合計

列レイアウト（R6 様式1_病棟票, ヘッダ5行）:
  col 0 : 医療機関コード（R6）
  col 2 : 都道府県コード
  col 3 : 二次医療圏コード
  col 4 : 二次医療圏名
  col 6 : 構想区域名
  col 11: 病棟コード
  col 18: 一般病床_許可病床
  col 22: 療養病床_許可病床
  データはR6から
"""
import json
import openpyxl
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6'
OUT = ROOT / 'data' / 'static' / 'medical_areas_national.json'
OUT_BACKUP = ROOT / 'data' / 'static' / 'medical_areas_national_R1_backup.json'

REGION_FILES = [
    'R6_様式1_北海道東北.xlsx',
    'R6_様式1_関東1.xlsx',
    'R6_様式1_関東2.xlsx',
    'R6_様式1_中部.xlsx',
    'R6_様式1_近畿.xlsx',
    'R6_様式1_中国四国.xlsx',
    'R6_様式1_九州沖縄.xlsx',
]

# 都道府県コード → 名称
PREF_CODE_TO_NAME = {
    '01': '北海道', '02': '青森県', '03': '岩手県', '04': '宮城県', '05': '秋田県',
    '06': '山形県', '07': '福島県', '08': '茨城県', '09': '栃木県', '10': '群馬県',
    '11': '埼玉県', '12': '千葉県', '13': '東京都', '14': '神奈川県', '15': '新潟県',
    '16': '富山県', '17': '石川県', '18': '福井県', '19': '山梨県', '20': '長野県',
    '21': '岐阜県', '22': '静岡県', '23': '愛知県', '24': '三重県', '25': '滋賀県',
    '26': '京都府', '27': '大阪府', '28': '兵庫県', '29': '奈良県', '30': '和歌山県',
    '31': '鳥取県', '32': '島根県', '33': '岡山県', '34': '広島県', '35': '山口県',
    '36': '徳島県', '37': '香川県', '38': '愛媛県', '39': '高知県', '40': '福岡県',
    '41': '佐賀県', '42': '長崎県', '43': '熊本県', '44': '大分県', '45': '宮崎県',
    '46': '鹿児島県', '47': '沖縄県',
}

DATA_START_ROW = 6  # 1-indexed; rows 1-5 are headers


def normalize_pref_code(code):
    if code is None:
        return None
    s = str(code).strip()
    if s.endswith('.0'):
        s = s[:-2]
    if s.isdigit():
        return s.zfill(2)
    return None


def safe_int(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s in ('', '-', '－', '‐', '未報告又はデータ不備'):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def process_region(filename, agg, hosp_set):
    print(f'  Processing {filename}...')
    wb = openpyxl.load_workbook(RAW / filename, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n_rows = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < DATA_START_ROW:
            continue
        if not row or len(row) < 23:
            continue
        hosp_code = row[0]
        pref_code = normalize_pref_code(row[2])
        area_code = row[3]
        area_name = row[4]
        ipan = safe_int(row[18])  # 一般病床_許可病床
        ryoyo = safe_int(row[22])  # 療養病床_許可病床
        if not pref_code or not area_name or not hosp_code:
            continue
        key = (pref_code, str(area_name).strip())
        agg[key]['wards'] += 1
        agg[key]['beds'] += ipan + ryoyo
        # 医療機関コードは文字列化してsetに記録
        hosp_set[key].add(str(hosp_code))
        n_rows += 1
    wb.close()
    print(f'    {n_rows} ward rows processed')


def main():
    print('=== R6 病床機能報告 ETL ===')
    if not RAW.exists():
        print(f'ERROR: {RAW} does not exist')
        return

    # バックアップ
    if OUT.exists() and not OUT_BACKUP.exists():
        import shutil
        shutil.copy(OUT, OUT_BACKUP)
        print(f'Backed up R1 → {OUT_BACKUP.name}')

    # 既存R1データ読み込み（比較用）
    with open(OUT, 'r', encoding='utf-8') as f:
        r1_data = json.load(f)
    r1_by_key = {(d['pref_code'], d['area']): d for d in r1_data}
    print(f'R1 baseline: {len(r1_data)} medical areas')

    # R6集計
    agg = defaultdict(lambda: {'wards': 0, 'beds': 0})
    hosp_set = defaultdict(set)
    for f in REGION_FILES:
        process_region(f, agg, hosp_set)

    # 結果整形
    result = []
    for (pref_code, area), v in agg.items():
        result.append({
            'pref_code': pref_code,
            'pref': PREF_CODE_TO_NAME.get(pref_code, '?'),
            'area': area,
            'hosp': len(hosp_set[(pref_code, area)]),
            'wards': v['wards'],
            'beds': v['beds'],
        })
    result.sort(key=lambda x: (x['pref_code'], x['area']))

    print(f'\nR6 result: {len(result)} medical areas')
    # 全国合計
    total_h = sum(r['hosp'] for r in result)
    total_w = sum(r['wards'] for r in result)
    total_b = sum(r['beds'] for r in result)
    print(f'  Total: hosp={total_h:,} wards={total_w:,} beds={total_b:,}')

    # R1合計と比較
    r1_total_h = sum(d.get('hosp', 0) for d in r1_data)
    r1_total_w = sum(d.get('wards', 0) for d in r1_data)
    r1_total_b = sum(d.get('beds', 0) for d in r1_data)
    print(f'  R1 (ref): hosp={r1_total_h:,} wards={r1_total_w:,} beds={r1_total_b:,}')
    print(f'  Δ:        hosp={total_h-r1_total_h:+,} wards={total_w-r1_total_w:+,} beds={total_b-r1_total_b:+,}')

    # サンプル比較（東京都・愛知県・京都府）
    print('\nSample comparison (R1 → R6):')
    for sample_pref in ['13', '23', '26', '47']:
        r1_areas = [d for d in r1_data if d['pref_code'] == sample_pref]
        r6_areas = [r for r in result if r['pref_code'] == sample_pref]
        pref = PREF_CODE_TO_NAME.get(sample_pref, '?')
        r1_b = sum(d.get('beds',0) for d in r1_areas)
        r6_b = sum(r['beds'] for r in r6_areas)
        r1_h = sum(d.get('hosp',0) for d in r1_areas)
        r6_h = sum(r['hosp'] for r in r6_areas)
        print(f'  {pref}: R1 areas={len(r1_areas)}/h={r1_h}/b={r1_b:,} → R6 areas={len(r6_areas)}/h={r6_h}/b={r6_b:,}')

    # 書き出し
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'\nWrote {OUT}')


if __name__ == '__main__':
    main()
