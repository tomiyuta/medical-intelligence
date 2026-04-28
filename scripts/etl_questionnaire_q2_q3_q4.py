#!/usr/bin/env python3
"""
ETL拡張: NDB特定健診質問票 Q2/Q3/Q4 を ndb_questionnaire.json に追加

データ出典: 厚労省 第10回NDBオープンデータ 特定健診 質問票（令和4年度）
- q2_pref.xlsx (001495779): Q2 血糖を下げる薬・インスリン使用 → 糖尿病Bridge
- q3_pref.xlsx (001495781): Q3 コレステロール・中性脂肪を下げる薬使用 → 循環器Bridge
- q4_pref.xlsx (001495783): Q4 脳卒中既往 → 脳血管Bridge

派生指標:
- diabetes_medication: Q2 「はい」/Q2合計 = 糖尿病薬・インスリン服用率
- lipid_medication:    Q3 「はい」/Q3合計 = 脂質異常症薬服用率
- stroke_history:      Q4 「はい」/Q4合計 = 脳卒中既往率

注: Q2/Q3/Q4 は県単位ファイル形式 (cols=20, Q6と同じ)
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_ndb_questionnaire'
OUT = ROOT / 'data' / 'static' / 'ndb_questionnaire.json'


def read_pref_questionnaire(filename):
    """県単位ファイル (cols=20) 読み取り
    cols: A=都道府県/B=回答/C-J=男年齢別+中計(col10)/K-R=女年齢別+中計(col18)
    """
    wb = openpyxl.load_workbook(RAW / filename, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pref_dist = {}
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = str(pref_cell).strip()
        if not last_pref: continue
        answer = ws.cell(r, 2).value
        if not answer: continue
        male_total = ws.cell(r, 10).value
        female_total = ws.cell(r, 18).value
        m = male_total if isinstance(male_total, (int, float)) else 0
        f = female_total if isinstance(female_total, (int, float)) else 0
        pref_dist.setdefault(last_pref, {}).setdefault(answer, 0)
        pref_dist[last_pref][answer] += m + f
    wb.close()
    return pref_dist


def compute_yes_rate(dist):
    total = sum(dist.values())
    if total <= 0: return None
    yes = dist.get('はい', 0)
    return round(yes / total * 1000) / 10


def main():
    existing = json.loads(OUT.read_text(encoding='utf-8'))
    
    print('Q2 (q2_pref.xlsx): 糖尿病薬・インスリン服用率 ETL...')
    q2 = {p: compute_yes_rate(d) for p, d in read_pref_questionnaire('q2_pref.xlsx').items()}
    
    print('Q3 (q3_pref.xlsx): 脂質異常症薬服用率 ETL...')
    q3 = {p: compute_yes_rate(d) for p, d in read_pref_questionnaire('q3_pref.xlsx').items()}
    
    print('Q4 (q4_pref.xlsx): 脳卒中既往率 ETL...')
    q4 = {p: compute_yes_rate(d) for p, d in read_pref_questionnaire('q4_pref.xlsx').items()}
    
    prefs = existing.get('prefectures', {})
    for pref in prefs:
        if pref in q2 and q2[pref] is not None: prefs[pref]['diabetes_medication'] = q2[pref]
        if pref in q3 and q3[pref] is not None: prefs[pref]['lipid_medication'] = q3[pref]
        if pref in q4 and q4[pref] is not None: prefs[pref]['stroke_history'] = q4[pref]
    
    if 'questions' not in existing: existing['questions'] = {}
    existing['questions'].update({
        'diabetes_medication': '現在、血糖を下げる薬又はインスリン注射を使用しているか (Q2, %)',
        'lipid_medication': '現在、コレステロールや中性脂肪を下げる薬を使用しているか (Q3, %)',
        'stroke_history': '医師から、脳卒中(脳出血、脳梗塞等)にかかっているといわれた、または治療を受けたことがあるか (Q4, %)',
    })
    
    OUT.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding='utf-8')
    
    print('\n=== 5県 sanity ===')
    for pref in ['東京都','大阪府','北海道','沖縄県','高知県']:
        p = prefs.get(pref, {})
        print(f'  {pref:<8} 糖尿病薬={p.get("diabetes_medication"):>5}% / 脂質薬={p.get("lipid_medication"):>5}% / 脳卒中既往={p.get("stroke_history"):>5}%')
    
    print('\n=== 47県分布 ===')
    for key, label in [('diabetes_medication','糖尿病薬'),('lipid_medication','脂質薬'),('stroke_history','脳卒中既往')]:
        vals = sorted([prefs[p].get(key) for p in prefs if prefs[p].get(key) is not None])
        if vals:
            n=len(vals)
            print(f'  {label}: n={n}, min={vals[0]:.1f}% / median={vals[n//2]:.1f}% / max={vals[-1]:.1f}%')


if __name__ == '__main__':
    main()
