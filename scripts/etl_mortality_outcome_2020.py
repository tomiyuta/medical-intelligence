#!/usr/bin/env python3
"""
ETL for mortality_outcome_2020.json (Phase 4-1 / P1-1 Layer 1)

Extracts both crude (粗死亡率) and age-adjusted (年齢調整死亡率) mortality
rates for 6 causes × 47 prefectures + national, by sex, from a single 2020
xlsx source.

Source:
  厚生労働省 令和5年度人口動態統計特殊報告
  令和2年都道府県別年齢調整死亡率の概況
  https://www.mhlw.go.jp/toukei/saikin/hw/jinkou/other/20sibou/xls/13.xlsx

Sheet structure (per design doc PHASE4_1_BRIDGE_OUTCOME_UI_DESIGN.md §14.2):
  参考１（１）: crude  / 全死因, 悪性新生物, 胃, 大腸, 肺, 糖尿病
  参考１（２）: crude  / 心疾患, 急性心筋梗塞, 脳血管疾患, 脳梗塞, 肺炎, COPD
  参考１（３）: crude  / 肝疾患, 腎不全, 老衰, 不慮の事故, 自殺
  参考２（１）: aam    / (同上 col layout)
  参考２（２）: aam    / (同上)
  参考２（３）: aam    / (同上)

Row layout (common):
  row 1: title
  row 3: cause headers (col 2,6,10,14,18,22)
  row 4: sex headers (男 col, 女 col+2)
  row 5: value type (率 col, 順位 col+1)
  row 6: 全国
  row 7-53: 47 prefectures
  row 54-55: notes

NOTE: This script is independent from etl_age_adjusted_mortality_2020.py
(Phase 3-1 Option B). Both ETLs and their JSONs coexist; this new JSON is
the canonical source for the Bridge UI Outcome cell (3-row display).
"""
import json
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / 'data' / 'raw_age_adjusted_mortality' / 'r2_age_adjusted_mortality.xlsx'
OUT_PATH = ROOT / 'data' / 'static' / 'mortality_outcome_2020.json'

SOURCE_URL = 'https://www.mhlw.go.jp/toukei/saikin/hw/jinkou/other/20sibou/xls/13.xlsx'
SOURCE_LABEL = '令和5年度人口動態統計特殊報告 令和2年都道府県別年齢調整死亡率の概況'

# 6 causes for Bridge Outcome
# (crude_sheet, crude_col), (aam_sheet, aam_col)
# col is the position of '男 率'; col+1=男順位, col+2=女率, col+3=女順位
CAUSE_LOC = {
    '悪性新生物': {
        'crude_sheet': '参考１（１）', 'crude_col': 6,
        'aam_sheet':   '参考２（１）', 'aam_col':   6,
    },
    '糖尿病': {
        'crude_sheet': '参考１（１）', 'crude_col': 22,
        'aam_sheet':   '参考２（１）', 'aam_col':   22,
    },
    '心疾患': {
        'crude_sheet': '参考１（２）', 'crude_col': 2,
        'aam_sheet':   '参考２（２）', 'aam_col':   2,
    },
    '脳血管疾患': {
        'crude_sheet': '参考１（２）', 'crude_col': 10,
        'aam_sheet':   '参考２（２）', 'aam_col':   10,
    },
    '肺炎': {
        'crude_sheet': '参考１（２）', 'crude_col': 18,
        'aam_sheet':   '参考２（２）', 'aam_col':   18,
    },
    '腎不全': {
        'crude_sheet': '参考１（３）', 'crude_col': 6,
        'aam_sheet':   '参考２（３）', 'aam_col':   6,
    },
}

# 47都道府県の正式名称マッピング
# xlsx 内では全角空白入りや異表記があるため、正規化キーで突合
PREF_NAMES = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県', '岐阜県',
    '静岡県', '愛知県', '三重県', '滋賀県', '京都府', '大阪府', '兵庫県',
    '奈良県', '和歌山県', '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県', '福岡県', '佐賀県', '長崎県',
    '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県',
]


def normalize_pref(s):
    """xlsx の都道府県セル値から全角空白等を除去して比較用キーを返す"""
    if s is None:
        return ''
    return re.sub(r'[\s\u3000]', '', str(s))


def find_pref_row(ws, target_pref):
    """正式名称 (例: '沖縄県') から xlsx 上の row 番号を引く。

    xlsx 内では '東　  京' '京　  都' '沖\u3000  縄' のように
    末尾の 都/府/県 が省略されかつ全角空白が入る変則表記がある。
    したがって正式名称の末尾1文字 (都/府/県) を1個だけ取り除いた
    短縮形でも突合できるようにする。
    """
    if target_pref.endswith(('都', '府', '県')):
        target_key_short = target_pref[:-1]
    else:
        target_key_short = target_pref  # '北海道' はそのまま
    for r in range(7, 54):  # row 7-53 = 47県
        cell_val = ws.cell(r, 1).value
        if cell_val is None:
            continue
        norm = normalize_pref(cell_val)
        # 完全一致 or 末尾の都道府県を取り除いた一致
        if norm == target_pref or norm == target_key_short:
            return r
    return None


def get_rate_rank(ws, row, col):
    """指定 (row, col) から rate と (col+1) から rank を取得"""
    rate = ws.cell(row, col).value
    rank = ws.cell(row, col + 1).value
    # rank は数値以外 (空欄, '①' 等) があり得るので int 変換は安全に
    rank_int = None
    if rank is not None and not isinstance(rank, str):
        try:
            rank_int = int(rank)
        except (ValueError, TypeError):
            rank_int = None
    elif isinstance(rank, str):
        # 全国行は '①' '\u3000' 等
        m = re.match(r'\s*(\d+)\s*', rank)
        if m:
            rank_int = int(m.group(1))
    return rate, rank_int


def extract_pref_cause(ws_crude, ws_aam, row_crude, row_aam, loc):
    """1県×1死因の {crude, age_adjusted} を抽出"""
    # crude
    c_male_rate, c_male_rank = get_rate_rank(ws_crude, row_crude, loc['crude_col'])
    c_female_rate, c_female_rank = get_rate_rank(ws_crude, row_crude, loc['crude_col'] + 2)
    # age_adjusted
    a_male_rate, a_male_rank = get_rate_rank(ws_aam, row_aam, loc['aam_col'])
    a_female_rate, a_female_rank = get_rate_rank(ws_aam, row_aam, loc['aam_col'] + 2)
    return {
        'crude': {
            'male':   {'rate': c_male_rate, 'rank': c_male_rank},
            'female': {'rate': c_female_rate, 'rank': c_female_rank},
        },
        'age_adjusted': {
            'male':   {'rate': a_male_rate, 'rank': a_male_rank},
            'female': {'rate': a_female_rate, 'rank': a_female_rank},
        },
    }


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f'xlsx not found: {XLSX_PATH}')

    print(f'[ETL] loading xlsx: {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    output = {
        'source': SOURCE_LABEL,
        'source_url': SOURCE_URL,
        'year': 2020,
        'unit': '人口10万対',
        'causes': list(CAUSE_LOC.keys()),
        'national': {},
        'prefectures': {},
        'notes': [
            'crude と age_adjusted は同じ2020年ソース (xlsx 参考1/2) から取得 — 比較可能',
            '2024年粗死亡率は別ファイル vital_stats_pref.json から参照',
            '2020年齢調整死亡率と2024年粗死亡率は直接比較しない (年次・補正有無が異なる)',
            'rank は1位最高(高値)、47位最低の順',
            '基準人口: 平成27年(2015年)モデル人口 (年齢調整時)',
        ],
    }

    # シート参照キャッシュ
    sheet_cache = {}
    def get_ws(name):
        if name not in sheet_cache:
            sheet_cache[name] = wb[name]
        return sheet_cache[name]

    # ─── 全国 (row 6) ───
    print('[ETL] extracting national (row 6)')
    for cause, loc in CAUSE_LOC.items():
        ws_c = get_ws(loc['crude_sheet'])
        ws_a = get_ws(loc['aam_sheet'])
        output['national'][cause] = extract_pref_cause(ws_c, ws_a, 6, 6, loc)

    # ─── 47都道府県 ───
    print('[ETL] extracting 47 prefectures')
    for pref in PREF_NAMES:
        output['prefectures'][pref] = {}
        for cause, loc in CAUSE_LOC.items():
            ws_c = get_ws(loc['crude_sheet'])
            ws_a = get_ws(loc['aam_sheet'])
            row_c = find_pref_row(ws_c, pref)
            row_a = find_pref_row(ws_a, pref)
            if row_c is None or row_a is None:
                print(f'  [WARN] row not found for {pref} in {loc["crude_sheet"]} or {loc["aam_sheet"]}')
                continue
            output['prefectures'][pref][cause] = extract_pref_cause(
                ws_c, ws_a, row_c, row_a, loc
            )

    # ─── 出力 ───
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # ─── サマリ ───
    n_pref = len(output['prefectures'])
    n_causes = len(output['causes'])
    has_national = bool(output['national'])
    print(f'[ETL] wrote: {OUT_PATH}')
    print(f'[ETL] prefectures: {n_pref}, causes: {n_causes}, national: {has_national}')

    # ─── fixture verification (Done 条件) ───
    print('\n[ETL] fixture verification:')
    checks = [
        ('沖縄県', '糖尿病', 'crude', 'male', 16.3),
        ('沖縄県', '糖尿病', 'age_adjusted', 'male', 20.8),
        ('沖縄県', '糖尿病', 'age_adjusted', 'female', 9.7),
        ('山口県', '肺炎', 'age_adjusted', 'male', 116.7),
        ('東京都', '糖尿病', 'crude', 'male', 10.8),
        ('秋田県', '脳血管疾患', 'age_adjusted', 'male', 124.1),
    ]
    all_pass = True
    for pref, cause, type_, sex, expected in checks:
        actual = output['prefectures'][pref][cause][type_][sex]['rate']
        ok = abs(actual - expected) < 0.05 if actual is not None else False
        status = 'OK' if ok else 'FAIL'
        if not ok:
            all_pass = False
        print(f'  [{status}] {pref} {cause} {type_} {sex}: expected {expected}, actual {actual}')

    if all_pass:
        print('\n[ETL] all fixture checks PASS')
    else:
        print('\n[ETL] some fixture checks FAILED')
        return 1
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
