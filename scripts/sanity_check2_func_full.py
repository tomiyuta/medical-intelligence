#!/usr/bin/env python3
"""
Sanity check #2 完全版: 全7地域の機能区分集計が圏域別合計と整合するか確認
"""
import openpyxl
import json
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
R6_DIR = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6'
JSON_PATH = ROOT / 'data' / 'static' / 'medical_areas_national.json'

REGION_FILES = [
    'R6_様式1_北海道東北.xlsx',
    'R6_様式1_関東1.xlsx',
    'R6_様式1_関東2.xlsx',
    'R6_様式1_中部.xlsx',
    'R6_様式1_近畿.xlsx',
    'R6_様式1_中国四国.xlsx',
    'R6_様式1_九州沖縄.xlsx',
]


def safe_int(v):
    if isinstance(v, (int, float)):
        return int(v)
    return 0


print('=== 全7地域機能区分集計（col 15: 2024年7月1日時点） ===\n')

func_total = defaultdict(lambda: {'wards': 0, 'beds': 0})
total_wards_all = 0
total_beds_all = 0

for fname in REGION_FILES:
    wb = openpyxl.load_workbook(R6_DIR / fname, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    region_wards = 0
    region_beds = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < 6:
            continue
        if not row or len(row) < 23:
            continue
        # pref_codeで実データ行か判定（"必須項目"等のヘッダ残骸を除外）
        pref_code = row[2]
        if pref_code is None:
            continue
        if isinstance(pref_code, str) and not pref_code.replace('.', '').isdigit():
            continue
        func = row[15]
        if not func:
            continue
        func_str = str(func).strip()
        beds = safe_int(row[18]) + safe_int(row[22])
        func_total[func_str]['wards'] += 1
        func_total[func_str]['beds'] += beds
        region_wards += 1
        region_beds += beds
    total_wards_all += region_wards
    total_beds_all += region_beds
    print(f'  {fname[3:-5]}: wards={region_wards} beds={region_beds:,}')
    wb.close()

print(f'\n  全国合計: wards={total_wards_all} beds={total_beds_all:,}')

print('\n=== 機能区分別集計 ===')
order = ['高度急性期', '急性期', '回復期', '慢性期']
core_beds = 0
core_wards = 0
for k in order:
    if k in func_total:
        v = func_total[k]
        print(f'  {k:<8}: wards={v["wards"]:>5} beds={v["beds"]:>9,}  ({v["beds"]/total_beds_all*100:.1f}%)')
        core_beds += v['beds']
        core_wards += v['wards']

print(f'  4機能合計: wards={core_wards} beds={core_beds:,}  ({core_beds/total_beds_all*100:.1f}%)')

print('\n  その他（休棟等）:')
for k, v in sorted(func_total.items(), key=lambda x: -x[1]['beds']):
    if k not in order:
        print(f'    {k[:30]:<30}: wards={v["wards"]:>5} beds={v["beds"]:>8,}')

# Compare to JSON aggregation (圏域別合計)
print('\n=== JSON集計との整合確認 ===')
r6_json = json.load(open(JSON_PATH))
json_total_beds = sum(d['beds'] for d in r6_json)
json_total_wards = sum(d['wards'] for d in r6_json)
print(f'  JSON 圏域別合計: wards={json_total_wards} beds={json_total_beds:,}')
print(f'  機能区分合計:    wards={total_wards_all} beds={total_beds_all:,}')
print(f'  差分:           wards={total_wards_all-json_total_wards:+} beds={total_beds_all-json_total_beds:+,}')
if abs(total_beds_all - json_total_beds) < 1000 and abs(total_wards_all - json_total_wards) < 50:
    print('  ✅ 圏域別合計と機能区分合計が概ね一致 → R6 ETL データ整合性OK')
else:
    print('  ⚠️ 差分大 → ETLロジック要再検討')
