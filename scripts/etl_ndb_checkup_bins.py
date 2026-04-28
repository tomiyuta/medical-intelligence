#!/usr/bin/env python3
"""
etl_ndb_checkup_bins.py — NDB第10回 特定健診 検査値階層別分布 ETL (Phase 1)

入力: data/raw_ndb_checkup/{metric}_pref.xlsx (BMI/HbA1c/SBP/LDL/UrineProtein)
出力:
- data/static/ndb_checkup_bins.json (階級分布保持、年齢階級別 county data)
- data/static/ndb_checkup_risk_rates.json (リスク率派生)

Excel構造 (NDB第10回 都道府県別性年齢階級別分布、cols=18):
  row 5+: A=都道府県(continued)/ B=検査値階層 / C-J=男年齢別7+中計 / K-R=女年齢別7+中計
  col 10 = 男中計, col 18 = 女中計, col 2 = 階層ラベル

Phase 1 リスク閾値:
- BMI: '25.0以上30.0未満' '30.0以上35.0未満' '35.0以上40.0未満' '40.0以上' → bmi_ge_25
- HbA1c: '6.5以上8.0未満' '8.0以上8.4未満' '8.4以上' → hba1c_ge_6_5
- 収縮期血圧: '140以上160未満' '160以上180未満' '180以上' → sbp_ge_140
- LDL: '140以上160未満' '160以上180未満' '180以上' → ldl_ge_140
- 尿蛋白: '＋' '＋＋' '＋＋＋' (- と ± は除外) → urine_protein_ge_1plus
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_ndb_checkup'
OUT_BINS = ROOT / 'data' / 'static' / 'ndb_checkup_bins.json'
OUT_RATES = ROOT / 'data' / 'static' / 'ndb_checkup_risk_rates.json'

# ファイル名 → metric キー
FILE_METRIC = {
    'BMI_pref.xlsx': 'BMI',
    'HbA1c_pref.xlsx': 'HbA1c',
    'SBP_pref.xlsx': '収縮期血圧',
    'LDL_pref.xlsx': 'LDL',
    'UrineProtein_pref.xlsx': '尿蛋白',
}

# リスク閾値: 該当する階層ラベルをセット
RISK_DEFS = {
    'BMI': {
        'risk_key': 'bmi_ge_25',
        'risk_label': 'BMI ≥25 (肥満)',
        'risk_bins': ['25.0以上30.0未満', '30.0以上35.0未満', '35.0以上40.0未満', '40.0以上'],
        'unit': 'kg/m²',
    },
    'HbA1c': {
        'risk_key': 'hba1c_ge_6_5',
        'risk_label': 'HbA1c ≥6.5% (糖尿病型)',
        'risk_bins': ['6.5以上8.0未満', '8.0以上8.4未満', '8.4以上'],
        'unit': '%',
    },
    '収縮期血圧': {
        'risk_key': 'sbp_ge_140',
        'risk_label': '収縮期血圧 ≥140 mmHg (高血圧)',
        'risk_bins': ['140以上160未満', '160以上180未満', '180以上'],
        'unit': 'mmHg',
    },
    'LDL': {
        'risk_key': 'ldl_ge_140',
        'risk_label': 'LDL ≥140 mg/dL (脂質異常症)',
        'risk_bins': ['140以上160未満', '160以上180未満', '180以上'],
        'unit': 'mg/dL',
    },
    '尿蛋白': {
        'risk_key': 'urine_protein_ge_1plus',
        'risk_label': '尿蛋白 1+以上 (CKDリスク)',
        'risk_bins': ['＋', '＋＋', '＋＋＋'],
        'unit': '定性',
    },
}


def normalize_bin(s):
    """階級ラベルを正規化 (空白除去, '～' → '-')"""
    if s is None: return None
    s = str(s).strip().replace('\u3000', '').replace('\xa0', '')
    return s


def num_or_zero(v):
    if isinstance(v, (int, float)): return int(v)
    return 0


def extract_bins(xlsx_path, metric):
    """Excel から (pref, bin_label, count_male, count_female) を抽出"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = str(pref_cell).strip()
        if not last_pref: continue
        bin_lbl = normalize_bin(ws.cell(r, 2).value)
        if not bin_lbl: continue
        male_total = num_or_zero(ws.cell(r, 10).value)
        female_total = num_or_zero(ws.cell(r, 18).value)
        if male_total == 0 and female_total == 0: continue
        records.append({
            'pref': last_pref,
            'metric': metric,
            'bin_label': bin_lbl,
            'count_male': male_total,
            'count_female': female_total,
            'count_total': male_total + female_total,
        })
    wb.close()
    return records


def compute_risk_rate(records, metric):
    """県別リスク率算出 (男女合算、全年齢)"""
    if metric not in RISK_DEFS: return {}
    risk_bins = set(RISK_DEFS[metric]['risk_bins'])
    by_pref = {}
    for r in records:
        by_pref.setdefault(r['pref'], []).append(r)
    rates = {}
    for pref, recs in by_pref.items():
        # 都道府県判別不可は除く
        if pref == '都道府県判別不可': continue
        total = sum(r['count_total'] for r in recs)
        if total == 0: continue
        ge_count = sum(r['count_total'] for r in recs if r['bin_label'] in risk_bins)
        rates[pref] = {
            'rate': round(ge_count/total*1000)/10,
            'denominator': total,
            'numerator': ge_count,
        }
    return rates


def main():
    if not RAW.exists():
        print(f'ERROR: {RAW} not found.')
        return
    
    all_bins = []
    all_rates = {}
    metric_summary = []
    
    for fname, metric in FILE_METRIC.items():
        xlsx = RAW / fname
        if not xlsx.exists():
            print(f'  MISSING: {fname}')
            continue
        print(f'  ETL: {fname} → metric={metric}')
        records = extract_bins(xlsx, metric)
        all_bins.extend(records)
        rates = compute_risk_rate(records, metric)
        rd = RISK_DEFS[metric]
        all_rates[rd['risk_key']] = {
            'metric': metric,
            'risk_label': rd['risk_label'],
            'risk_bins': rd['risk_bins'],
            'unit': rd['unit'],
            'by_pref': rates,
        }
        # サマリ
        if rates:
            vals = sorted([v['rate'] for v in rates.values()])
            metric_summary.append((metric, rd['risk_key'], len(rates), vals[0], vals[len(vals)//2], vals[-1]))
    
    # 保存
    with open(OUT_BINS, 'w', encoding='utf-8') as f:
        json.dump({
            'source': 'NDB第10回オープンデータ 特定健診 検査値階層別分布 (令和4年度=2022)',
            'source_url': 'https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000177221_00016.html',
            'data': all_bins,
        }, f, ensure_ascii=False, indent=2)
    print(f'\n保存: {OUT_BINS} ({len(all_bins)} records)')
    
    with open(OUT_RATES, 'w', encoding='utf-8') as f:
        json.dump({
            'source': 'NDB第10回オープンデータ 特定健診 リスク該当者率 (派生)',
            'method': '男女・全年齢合計から各リスク閾値超過の階級人数 / 該当県の全階級人数',
            'risk_rates': all_rates,
        }, f, ensure_ascii=False, indent=2)
    print(f'保存: {OUT_RATES} ({len(all_rates)} risk metrics)')
    
    # サマリ
    print('\n=== Phase 1 リスク率算出サマリ ===')
    print(f'{"metric":<10}{"risk_key":<25}{"n":>5}{"min":>8}{"median":>10}{"max":>8}')
    for metric, rk, n, lo, med, hi in metric_summary:
        print(f'{metric:<10}{rk:<25}{n:>5}{lo:>7.1f}%{med:>9.1f}%{hi:>7.1f}%')
    
    # 5県 sanity
    print('\n=== 5県 sanity (リスク率) ===')
    target_prefs = ['東京都','大阪府','北海道','沖縄県','高知県']
    headers = ['BMI≥25', 'HbA1c≥6.5', 'SBP≥140', 'LDL≥140', '尿蛋白1+']
    keys = ['bmi_ge_25', 'hba1c_ge_6_5', 'sbp_ge_140', 'ldl_ge_140', 'urine_protein_ge_1plus']
    print(f'{"県":<8}' + ''.join(f'{h:>11}' for h in headers))
    for pref in target_prefs:
        row = [f'{pref:<8}']
        for k in keys:
            r = all_rates.get(k, {}).get('by_pref', {}).get(pref, {}).get('rate')
            row.append(f'{r:>10.1f}%' if r is not None else f'{"-":>11}')
        print(''.join(row))


if __name__ == '__main__':
    main()
