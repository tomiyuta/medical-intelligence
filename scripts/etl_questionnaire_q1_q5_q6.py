#!/usr/bin/env python3
"""
phase2 ETL拡張: NDB特定健診質問票 Q1/Q5/Q6 を ndb_questionnaire.json に追加

データ出典: 厚労省 第10回NDBオープンデータ 特定健診 質問票（令和4年度）
- q1_pref.xlsx: Q1 血圧降下剤の服用 (はい/いいえ) ← 循環器・脳血管Bridge用
- 786.xlsx:    Q5 心臓病既往 (はい/いいえ)        ← 循環器Bridge用
- 788.xlsx:    Q6 慢性腎臓病/腎不全/人工透析既往  ← 腎疾患Bridge用 (県単位ファイル使用)

派生指標:
- hypertension_med: Q1 「はい」/Q1合計 = 高血圧薬服用率
- heart_disease:    Q5 「はい」/Q5合計 = 心臓病既往率
- ckd_history:      Q6 「はい」/Q6合計 = CKD既往率

注:
- Q1/Q5 は 二次医療圏ファイル (cols=22) → 県単位で aggregate
- Q6 は 県単位ファイル (cols=20) → そのまま使用

未取得 (user取得依頼中):
- Q2 (糖尿病薬・インスリン服用) — 循環器・糖尿病Bridge強化
- Q3 (脂質異常症薬服用) — 循環器Bridge強化
- Q4 (脳卒中既往) — 脳血管Bridge強化
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_ndb_questionnaire'
OUT = ROOT / 'data' / 'static' / 'ndb_questionnaire.json'


def read_q1_or_q5(filename):
    """二次医療圏ファイル形式 (q1_pref.xlsx / 786.xlsx) を県単位に aggregate
    cols: A=都道府県/B=圏番号/C=圏名/D=回答/E-L=男年齢別+中計/M-T=女年齢別+中計
    男合計=col 12 (中計), 女合計=col 20 (中計)
    """
    wb = openpyxl.load_workbook(RAW / filename, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pref_dist = {}
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = pref_cell
        if not last_pref: continue
        answer = ws.cell(r, 4).value
        if not answer: continue
        male_total = ws.cell(r, 12).value
        female_total = ws.cell(r, 20).value
        m = male_total if isinstance(male_total, (int, float)) else 0
        f = female_total if isinstance(female_total, (int, float)) else 0
        pref_dist.setdefault(last_pref, {}).setdefault(answer, 0)
        pref_dist[last_pref][answer] += m + f
    wb.close()
    return pref_dist


def read_q6(filename):
    """県単位ファイル形式 (788.xlsx)
    cols: A=都道府県/B=回答/C-J=男年齢別+中計/K-R=女年齢別+中計
    男合計=col 10 (中計), 女合計=col 18 (中計)
    """
    wb = openpyxl.load_workbook(RAW / filename, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pref_dist = {}
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = pref_cell
        if not last_pref: continue
        answer = ws.cell(r, 2).value
        if not answer: continue
        male_total = ws.cell(r, 10).value
        female_total = ws.cell(r, 18).value
        m = male_total if isinstance(male_total, (int, float)) else 0
        f = female_total if isinstance(female_total, (int, float)) else 0
        pref_dist.setdefault(last_pref, {}).setdefault(answer, 0)
        pref_dist[last_pref][answer] += m + f
    wb.close()
    return pref_dist


def compute_yes_rate(dist):
    total = sum(dist.values())
    if total <= 0: return None
    yes = dist.get('はい', 0)
    return round(yes / total * 1000) / 10


def main():
    # 既存 JSON ロード
    existing = json.loads(OUT.read_text(encoding='utf-8'))
    
    # Q1: 血圧降下剤服用
    print('Q1 (q1_pref.xlsx): 高血圧薬服用率 ETL...')
    q1_data = read_q1_or_q5('q1_pref.xlsx')
    q1_rates = {pref: compute_yes_rate(d) for pref, d in q1_data.items()}
    
    # Q5: 心臓病既往
    print('Q5 (786.xlsx): 心臓病既往率 ETL...')
    q5_data = read_q1_or_q5('786.xlsx')
    q5_rates = {pref: compute_yes_rate(d) for pref, d in q5_data.items()}
    
    # Q6: CKD既往
    print('Q6 (788.xlsx): CKD既往率 ETL...')
    q6_data = read_q6('788.xlsx')
    q6_rates = {pref: compute_yes_rate(d) for pref, d in q6_data.items()}
    
    # 既存 prefectures に追加
    prefs = existing.get('prefectures', {})
    added_count = 0
    for pref in prefs:
        if pref in q1_rates and q1_rates[pref] is not None:
            prefs[pref]['hypertension_med'] = q1_rates[pref]
            added_count += 1
        if pref in q5_rates and q5_rates[pref] is not None:
            prefs[pref]['heart_disease'] = q5_rates[pref]
        if pref in q6_rates and q6_rates[pref] is not None:
            prefs[pref]['ckd_history'] = q6_rates[pref]
    
    # questions メタデータ更新
    if 'questions' not in existing:
        existing['questions'] = {}
    existing['questions'].update({
        'hypertension_med': '現在、血圧を下げる薬を使用しているか (Q1, %)',
        'heart_disease': '医師から心臓病(狭心症・心筋梗塞等)にかかっているといわれた、または治療を受けたことがあるか (Q5, %)',
        'ckd_history': '医師から慢性腎臓病や腎不全にかかっているといわれた、または治療(人工透析)を受けたことがあるか (Q6, %)',
    })
    
    OUT.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'\n更新完了: {added_count} 県に Q1/Q5/Q6 追加')
    
    # 5県 sanity
    print('\n=== 5県 sanity (高血圧薬服用率/心臓病既往率/CKD既往率) ===')
    for pref in ['東京都','大阪府','北海道','沖縄県','高知県']:
        p = prefs.get(pref, {})
        print(f'  {pref:<8} 高血圧薬={p.get("hypertension_med"):>5}% / 心臓病={p.get("heart_disease"):>5}% / CKD={p.get("ckd_history"):>5}%')
    
    # 47県分布
    print('\n=== 47県分布 ===')
    for key, label in [('hypertension_med', '高血圧薬服用率'), ('heart_disease', '心臓病既往率'), ('ckd_history', 'CKD既往率')]:
        vals = sorted([prefs[p].get(key) for p in prefs if prefs[p].get(key) is not None])
        if vals:
            n = len(vals)
            print(f'  {label}: n={n}, min={vals[0]:.1f}% / median={vals[n//2]:.1f}% / mean={sum(vals)/n:.1f}% / max={vals[-1]:.1f}%')


if __name__ == '__main__':
    main()
