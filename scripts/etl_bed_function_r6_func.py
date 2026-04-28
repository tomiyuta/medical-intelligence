#!/usr/bin/env python3
"""
phase7 ETL: R6 病床機能報告の col 15 (機能区分) を都道府県別×機能区分別で集計

出力: data/static/bed_function_by_pref.json

入力: 既ダウンロード済 data/raw/source/06_病床機能報告/data_R6/R6_様式1_*.xlsx
列レイアウト (R6 様式1_病棟票, ヘッダ5行):
  col 2 : 都道府県コード
  col 3 : 二次医療圏コード
  col 15: 保有する病棟と機能区分の選択状況（2024（令和6）年7月1日時点）
  col 18: 一般病床_許可病床
  col 22: 療養病床_許可病床
"""
import json
import openpyxl
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6'
OUT = ROOT / 'data' / 'static' / 'bed_function_by_pref.json'

REGION_FILES = [
    'R6_様式1_北海道東北.xlsx',
    'R6_様式1_関東1.xlsx',
    'R6_様式1_関東2.xlsx',
    'R6_様式1_中部.xlsx',
    'R6_様式1_近畿.xlsx',
    'R6_様式1_中国四国.xlsx',
    'R6_様式1_九州沖縄.xlsx',
]

PREF_CODE_TO_NAME = {
    '01': '北海道', '02': '青森県', '03': '岩手県', '04': '宮城県', '05': '秋田県',
    '06': '山形県', '07': '福島県', '08': '茨城県', '09': '栃木県', '10': '群馬県',
    '11': '埼玉県', '12': '千葉県', '13': '東京都', '14': '神奈川県', '15': '新潟県',
    '16': '富山県', '17': '石川県', '18': '福井県', '19': '山梨県', '20': '長野県',
    '21': '岐阜県', '22': '静岡県', '23': '愛知県', '24': '三重県', '25': '滋賀県',
    '26': '京都府', '27': '大阪府', '28': '兵庫県', '29': '奈良県', '30': '和歌山県',
    '31': '鳥取県', '32': '島根県', '33': '岡山県', '34': '広島県', '35': '山口県',
    '36': '徳島県', '37': '香川県', '38': '愛媛県', '39': '高知県', '40': '福岡県',
    '41': '佐賀県', '42': '長崎県', '43': '熊本県', '44': '大分県', '45': '宮崎県',
    '46': '鹿児島県', '47': '沖縄県',
}

# 機能区分の正規化 (UIで扱いやすい4機能 + 休棟2 + その他)
ACTIVE_4 = {'高度急性期', '急性期', '回復期', '慢性期'}
INACTIVE = {'休棟中（今後再開する予定）': '休棟(再開)', '休棟中（今後廃止する予定）': '休棟(廃止)'}

DATA_START_ROW = 6


def normalize_pref_code(code):
    if code is None:
        return None
    s = str(code).strip()
    if s.endswith('.0'):
        s = s[:-2]
    if s.isdigit():
        return s.zfill(2)
    return None


def safe_int(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s in ('', '-', '－', '‐', '未報告又はデータ不備'):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def classify_function(val):
    """機能区分文字列を正規化キーに変換。未分類は None。"""
    if not val:
        return None
    s = str(val).strip()
    if s in ACTIVE_4:
        return s
    if s in INACTIVE:
        return INACTIVE[s]
    return None


def main():
    print('=== R6 機能区分 ETL (都道府県別×機能別) ===')
    if not RAW.exists():
        print(f'ERROR: {RAW} does not exist')
        return

    # pref → func → {wards, beds}
    agg = defaultdict(lambda: defaultdict(lambda: {'wards': 0, 'beds': 0}))
    skipped_func = defaultdict(int)

    for fname in REGION_FILES:
        print(f'  Processing {fname}...')
        wb = openpyxl.load_workbook(RAW / fname, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        n_processed, n_classified = 0, 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx < DATA_START_ROW:
                continue
            if not row or len(row) < 23:
                continue
            pref_code = normalize_pref_code(row[2])
            func_raw = row[15]
            ipan = safe_int(row[18])
            ryoyo = safe_int(row[22])
            if not pref_code:
                continue
            n_processed += 1
            func = classify_function(func_raw)
            if func is None:
                if func_raw:
                    skipped_func[str(func_raw).strip()] += 1
                continue
            n_classified += 1
            agg[pref_code][func]['wards'] += 1
            agg[pref_code][func]['beds'] += ipan + ryoyo
        wb.close()
        print(f'    {n_processed} rows processed, {n_classified} classified')

    if skipped_func:
        print(f'\nSkipped non-classifiable func values: {dict(skipped_func)}')

    # 全国集計
    national = defaultdict(lambda: {'wards': 0, 'beds': 0})
    for pref_code, funcs in agg.items():
        for func, v in funcs.items():
            national[func]['wards'] += v['wards']
            national[func]['beds'] += v['beds']

    # 出力整形
    categories = ['高度急性期', '急性期', '回復期', '慢性期', '休棟(再開)', '休棟(廃止)']
    output = {
        'source': '厚労省 令和6年度病床機能報告 (2024年7月1日時点)',
        'published': '2025-09-30',
        'note': '機能区分は施設の自己申告に基づく。「2024年7月1日時点の機能」を採用。地域医療構想の評価指標として使用。',
        'categories': categories,
        'national': {c: dict(national.get(c, {'wards': 0, 'beds': 0})) for c in categories},
        'prefectures': {},
    }
    nat_total_beds = sum(national[c]['beds'] for c in categories)
    output['national']['総床数'] = nat_total_beds
    print(f'\n全国総床数: {nat_total_beds:,}')
    for c in categories:
        v = national[c]
        share = v['beds'] / nat_total_beds * 100 if nat_total_beds > 0 else 0
        print(f'  {c:<10}: 病棟={v["wards"]:>5,} 床={v["beds"]:>9,} ({share:>5.1f}%)')

    # 都道府県別
    for pref_code, funcs in sorted(agg.items()):
        pref = PREF_CODE_TO_NAME.get(pref_code, '?')
        pref_total = sum(funcs[c]['beds'] for c in categories)
        output['prefectures'][pref] = {
            **{c: dict(funcs.get(c, {'wards': 0, 'beds': 0})) for c in categories},
            '総床数': pref_total,
        }

    # サンプル表示
    print('\n=== サンプル都道府県 ===')
    for sample_pref in ['東京都', '愛知県', '京都府', '秋田県', '高知県', '沖縄県']:
        d = output['prefectures'].get(sample_pref)
        if not d:
            continue
        tot = d['総床数']
        shares = []
        for c in ['高度急性期', '急性期', '回復期', '慢性期']:
            s = d[c]['beds'] / tot * 100 if tot > 0 else 0
            shares.append(f'{c}={s:.1f}%')
        print(f'  {sample_pref}: 総={tot:,} {" / ".join(shares)}')

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'\nWrote {OUT}')


if __name__ == '__main__':
    main()
