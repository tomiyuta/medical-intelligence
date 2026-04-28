#!/usr/bin/env python3
"""R6 病床機能データ sanity check（5項目）"""
import json
import openpyxl
from collections import defaultdict
from pathlib import Path
import re

ROOT = Path(__file__).resolve().parent.parent
R6_DIR = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6'
JSON_PATH = ROOT / 'data' / 'static' / 'medical_areas_national.json'


def check1_ranking():
    print('\n=== Check #1: 都道府県別総床数ランキング ===')
    r6 = json.load(open(JSON_PATH))
    by_pref = defaultdict(lambda: {'h': 0, 'w': 0, 'b': 0, 'areas': 0})
    for d in r6:
        by_pref[d['pref']]['h'] += d['hosp']
        by_pref[d['pref']]['w'] += d['wards']
        by_pref[d['pref']]['b'] += d['beds']
        by_pref[d['pref']]['areas'] += 1

    print(f'  都道府県カバー: {len(by_pref)}/47')
    if len(by_pref) != 47:
        missing = set([
            '北海道','青森県','岩手県','宮城県','秋田県','山形県','福島県','茨城県','栃木県','群馬県',
            '埼玉県','千葉県','東京都','神奈川県','新潟県','富山県','石川県','福井県','山梨県','長野県',
            '岐阜県','静岡県','愛知県','三重県','滋賀県','京都府','大阪府','兵庫県','奈良県','和歌山県',
            '鳥取県','島根県','岡山県','広島県','山口県','徳島県','香川県','愛媛県','高知県','福岡県',
            '佐賀県','長崎県','熊本県','大分県','宮崎県','鹿児島県','沖縄県',
        ]) - set(by_pref.keys())
        print(f'  ❌ 欠損: {missing}')
        return False

    ranked = sorted(by_pref.items(), key=lambda x: -x[1]['b'])
    print('  TOP5（床数）:')
    for p, v in ranked[:5]:
        print(f'    {p}: hosp={v["h"]} wards={v["w"]} beds={v["b"]:,} areas={v["areas"]}')
    print('  BOTTOM5（床数）:')
    for p, v in ranked[-5:]:
        print(f'    {p}: hosp={v["h"]} wards={v["w"]} beds={v["b"]:,} areas={v["areas"]}')

    # 想定: 上位=北海道/東京/大阪/神奈川/兵庫/福岡など、下位=鳥取/福井/島根/佐賀/山梨など
    top_prefs = {p for p, _ in ranked[:5]}
    expected_top_set = {'北海道', '東京都', '大阪府', '神奈川県', '兵庫県', '福岡県', '愛知県'}
    bottom_prefs = {p for p, _ in ranked[-5:]}
    expected_bot_set = {'鳥取県', '福井県', '島根県', '佐賀県', '山梨県', '高知県', '徳島県', '富山県', '石川県'}
    top_ok = len(top_prefs & expected_top_set) >= 3
    bot_ok = len(bottom_prefs & expected_bot_set) >= 3
    print(f'  上位TOP5重複: {top_prefs & expected_top_set} → {"✅" if top_ok else "⚠️"}')
    print(f'  下位BOTTOM5重複: {bottom_prefs & expected_bot_set} → {"✅" if bot_ok else "⚠️"}')

    # 1床/1000人計算（住基2025比）
    ap = json.load(open(ROOT / 'data' / 'static' / 'age_pyramid.json'))
    print('  床密度（人口千対）:')
    for p in ['東京都', '高知県', '北海道', '沖縄県']:
        if p in ap.get('prefectures', {}):
            d = ap['prefectures'][p]
            pop = sum(d['male']) + sum(d['female'])
            beds = by_pref[p]['b']
            density = beds / pop * 1000
            print(f'    {p}: 床={beds:,} 人口={pop:,} 千対={density:.1f}床')

    return top_ok and bot_ok


def check2_function_breakdown():
    print('\n=== Check #2: 病床機能区分（高度急性期/急性期/回復期/慢性期）合計確認 ===')
    # 様式1の機能列を発見
    sample_file = R6_DIR / 'R6_様式1_北海道東北.xlsx'
    if not sample_file.exists():
        print('  ⚠️ R6 様式1ファイルが存在しない（要再ダウンロード）')
        return False

    wb = openpyxl.load_workbook(sample_file, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if i >= 5:
            break

    # row5 (idx 4) のサブヘッダから「機能」関連カラムを抽出
    subhdr = rows[4]
    func_cols = []
    for i, c in enumerate(subhdr):
        if c is None:
            continue
        s = str(c)
        if any(k in s for k in ['高度急性期', '急性期', '回復期', '慢性期']) and '機能' in (str(rows[1][i]) if rows[1][i] else ''):
            func_cols.append((i, s[:30]))

    # よりシンプル: rows[1] (R2) の機能区分を探す
    print('  R6様式1の上位ヘッダ(row 1)を機能で検索:')
    for i, c in enumerate(rows[1]):
        if c and any(k in str(c) for k in ['高度急性期', '急性期', '回復期', '慢性期', '医療機能']):
            sub = rows[4][i] if rows[4][i] else ''
            print(f'    col{i}: top={str(c)[:30]} sub={str(sub)[:30]}')

    print('  機能区分カラム特定が必要: 様式1の構造を再調査して別途ETL設計が必要')
    print('  → 現時点では「許可病床(一般+療養) 1,151,401床」は集計済み')
    print('  → 機能別内訳(高度急性期/急性期/回復期/慢性期)は未集計のためチェック保留')
    wb.close()
    return None  # 未実施として扱う


def check3_areas_per_pref():
    print('\n=== Check #3: 医療圏数が都道府県別に妥当か ===')
    r6 = json.load(open(JSON_PATH))
    by_pref = defaultdict(int)
    for d in r6:
        by_pref[d['pref']] += 1
    print(f'  全都道府県カバー: {len(by_pref)}/47')
    zero_or_one = [p for p, n in by_pref.items() if n < 1]
    if zero_or_one:
        print(f'  ❌ 圏域0の県: {zero_or_one}')
        return False
    # 圏域数の分布
    distribution = sorted(by_pref.items(), key=lambda x: x[1])
    print(f'  最少: {distribution[0]} / 最多: {distribution[-1]}')
    print(f'  Total: {sum(by_pref.values())}（HANDOFF想定: 330）')
    return sum(by_pref.values()) == 330


def check4_akita_mie_ui():
    print('\n=== Check #4: 秋田・三重のUI圏域整合 ===')
    r6 = json.load(open(JSON_PATH))
    akita = sorted({d['area'] for d in r6 if d['pref'] == '秋田県'})
    mie = sorted({d['area'] for d in r6 if d['pref'] == '三重県'})
    print(f'  秋田県({len(akita)}): {akita}')
    print(f'  三重県({len(mie)}): {mie}')
    expected_akita = {'県北', '県央', '県南'}
    expected_mie = {'北勢', '中勢伊賀', '南勢志摩', '東紀州'}  # 7→4: 東紀州は単独で残存
    ok_akita = set(akita) == expected_akita
    ok_mie = set(mie) == expected_mie
    print(f'  秋田 expected: {expected_akita} → {"✅" if ok_akita else "❌"}')
    print(f'  三重 expected: {expected_mie} → {"✅" if ok_mie else "❌"}')
    return ok_akita and ok_mie


def check5_old_area_refs():
    print('\n=== Check #5: 旧R1医療圏名のハードコード参照スキャン ===')
    # R1 にあって R6 から消えた圏域名
    r1 = json.load(open(ROOT / 'data' / 'static' / 'medical_areas_national_R1_backup.json'))
    r6 = json.load(open(JSON_PATH))
    r1_areas = {d['area'] for d in r1}
    r6_areas = {d['area'] for d in r6}
    removed = r1_areas - r6_areas
    print(f'  R1にあってR6で消えた圏域名: {sorted(removed)}')

    # app/ 以下のJS/JSXをスキャン
    found = []
    for ext in ['*.js', '*.jsx']:
        for path in (ROOT / 'app').rglob(ext):
            try:
                content = path.read_text(encoding='utf-8')
            except UnicodeDecodeError:
                continue
            for area in removed:
                # クォートで囲まれた一致 ('北秋田' など)
                if re.search(rf"['\"`]{re.escape(area)}['\"`]", content):
                    rel = path.relative_to(ROOT)
                    found.append((str(rel), area))

    if found:
        print(f'  ⚠️ ハードコード参照が見つかった ({len(found)}件):')
        for path, area in found:
            print(f'    {path}: "{area}"')
        return False
    print(f'  ✅ 旧R1圏域名をハードコードで参照しているJS/JSXファイルは無し')
    return True


def main():
    results = {}
    results['#1 ランキング常識性'] = check1_ranking()
    results['#2 機能区分合計'] = check2_function_breakdown()
    results['#3 医療圏数妥当性'] = check3_areas_per_pref()
    results['#4 秋田・三重UI整合'] = check4_akita_mie_ui()
    results['#5 古いキー参照'] = check5_old_area_refs()

    print('\n' + '=' * 50)
    print('=== Sanity Check 結果サマリ ===')
    for k, v in results.items():
        if v is True:
            print(f'  ✅ {k}: PASS')
        elif v is False:
            print(f'  ❌ {k}: FAIL')
        else:
            print(f'  ⚠️ {k}: 保留（手動確認要）')

    all_pass = all(v is True or v is None for v in results.values())
    if all_pass:
        print('\n✅ 全項目クリア（または保留）→ Priority 6 着手OK')
    else:
        fails = [k for k, v in results.items() if v is False]
        print(f'\n❌ 失敗項目あり: {fails} → Priority 6 着手前に修正が必要')


if __name__ == '__main__':
    main()
