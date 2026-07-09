#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tier B ETL: DPC症例数・平均在院日数の推移（カルテ#38 圏 / #40 都道府県別）。
DPC退院患者調査「在院日数の状況」(施設別・5年ローリング) を 2018版(2016-2018) + 2023版(2019-2023)で
2016-2023を充足。施設→二次医療圏/都道府県 を通番→市町村番号→圏 で集約し、件数加重の平均在院日数と
症例数を年次推移で算出。カルテ#38(山城南)/#40(都道府県別)と数値一致。

出力: data/static/dpc_los_trend.json
"""
import csv
import re
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
D23 = ROOT / 'data' / 'raw' / 'source' / 'DPC退院患者調査_R5'
D18 = ROOT / 'data' / 'raw' / 'source' / 'DPC退院患者調査_2018'
SHISETSU = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6' / 'R6_施設票.xlsx'
MASTER = ROOT / 'data' / 'hsa' / 'area_master.json'
OUT = ROOT / 'data' / 'static' / 'dpc_los_trend.json'

# 在院日数ファイルの年→(件数列, 平均値列=件数列+1)
YCOLS_18 = {2016: 19, 2017: 27, 2018: 35}
YCOLS_23 = {2019: 3, 2020: 11, 2021: 19, 2022: 27, 2023: 35}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if s in ('', '-', '－', '*', '…'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def build_muni2hsa():
    master = {(m['pref_code'], m['siteArea']): m['code']
              for m in json.load(open(MASTER, encoding='utf-8'))['areas'] if m.get('siteArea')}
    wb = openpyxl.load_workbook(SHISETSU, read_only=True, data_only=True)
    ws = wb['Sheet1']
    m2h = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        mc, pc, an = row[8], row[3], row[5]
        if mc and pc and an:
            mc = str(mc).strip().split('.')[0].zfill(5)
            pc = str(pc).strip().split('.')[0].zfill(2)
            h = master.get((pc, str(an).strip()))
            if h:
                m2h[mc] = h
    wb.close()
    return m2h


def tsuban_map(fac_xlsx):
    """通番 → (市町村番号, DPC対象病院か)。病院類型(col5)に'DPC'を含めばDPC対象。"""
    wb = openpyxl.load_workbook(fac_xlsx, read_only=True, data_only=True)
    ws = wb['施設概要表']
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[1]
        if t is None:
            continue
        t = str(t).strip().split('.')[0]
        mc = str(row[2]).strip().split('.')[0].zfill(5) if row[2] is not None else ''
        ruikei = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        out[t] = (mc, 'DPC' in ruikei)
    wb.close()
    return out


def accumulate(los_xlsx, fac_xlsx, ycols, m2h, area_acc, pref_acc, nat_acc):
    t2muni = tsuban_map(fac_xlsx)
    wb = openpyxl.load_workbook(los_xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=4, values_only=True):
        t = row[1]
        if t is None:
            continue
        t = str(t).strip().split('.')[0]
        muni, is_dpc = t2muni.get(t, ('', False))
        hsa = m2h.get(muni)
        pref = muni[:2] if muni else ''
        for y, c in ycols.items():
            k = num(row[c] if len(row) > c else None)
            los = num(row[c + 1] if len(row) > c + 1 else None)
            if not k or los is None:
                continue
            # 圏(#38): DPC対象病院のみ
            if hsa and is_dpc:
                a = area_acc.setdefault(hsa, {}).setdefault(y, [0.0, 0.0])
                a[0] += k; a[1] += k * los
            # 都道府県/全国(#40): DPC参加病院すべて(対象+準備+出来高)
            if pref:
                p = pref_acc.setdefault(pref, {}).setdefault(y, [0.0, 0.0])
                p[0] += k; p[1] += k * los
            n = nat_acc.setdefault(y, [0.0, 0.0])
            n[0] += k; n[1] += k * los
    wb.close()


def main():
    m2h = build_muni2hsa()
    master = {m['code']: m for m in json.load(open(MASTER, encoding='utf-8'))['areas']}
    area_acc, pref_acc, nat_acc = {}, {}, {}
    accumulate(D18 / '在院日数の状況.xlsx', D18 / '施設概要表.xlsx', YCOLS_18, m2h, area_acc, pref_acc, nat_acc)
    accumulate(D23 / '在院日数の状況.xlsx', D23 / '施設概要表.xlsx', YCOLS_23, m2h, area_acc, pref_acc, nat_acc)

    def fin(acc):  # {y:[件数, 件数×los]} → {y:{kensu, los}}
        return {str(y): {'kensu': round(v[0]), 'los': round(v[1] / v[0], 1) if v[0] else None} for y, v in acc.items()}

    areas = {}
    for code, yd in area_acc.items():
        m = master.get(code, {})
        areas[code] = {'pref': m.get('pref', ''), 'area': m.get('area', ''), 'years': fin(yd)}
    # pref: コード→名
    prefname = {}
    for m in master.values():
        prefname[m['code'][:2]] = m['pref']
    prefs = {prefname.get(pc, pc): {str(y): (round(v[1] / v[0], 2) if v[0] else None) for y, v in yd.items()} for pc, yd in pref_acc.items()}
    national = {str(y): (round(v[1] / v[0], 2) if v[0] else None) for y, v in nat_acc.items()}

    payload = {
        'source': '厚生労働省「DPC導入の影響評価に係る調査（退院患者調査）」在院日数の状況（2018版＋2023版）',
        'note': 'DPC症例数と平均在院日数の年度推移(2016-2023)。施設→二次医療圏/都道府県を通番→市町村で集約、'
                '平均在院日数は件数加重。カルテ#38/#40と数値一致。',
        'years': list(range(2016, 2024)),
        'areaCount': len(areas), 'areas': areas, 'prefs': prefs, 'national': national,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
    print(f"[dpc-los] areas={len(areas)} prefs={len(prefs)} out={OUT.name}")
    ya = areas.get('2606')
    if ya:
        print('[dpc-los] 山城南 症例数:', {y: ya['years'].get(str(y), {}).get('kensu') for y in (2016, 2018, 2023)}, '(カルテ 4431/4525/4147)')
        print('  平均在院日数:', {y: ya['years'].get(str(y), {}).get('los') for y in (2016, 2018, 2023)}, '(10.6/10.5/10.0)')
    print('[dpc-los] 全国 平均在院日数:', {y: national.get(str(y)) for y in (2016, 2018, 2023)}, '(カルテ 12.43/12.05/11.72)')
    print('[dpc-los] 京都府:', {y: prefs.get('京都府', {}).get(y) for y in (2016, 2018, 2023)}, '(12.16/11.71/11.22)')


if __name__ == '__main__':
    main()
