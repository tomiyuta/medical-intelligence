#!/usr/bin/env python3
"""
Phase 3-1 Option B: 年齢調整死亡率 2020 ETL

入力: data/raw_age_adjusted_mortality/r2_age_adjusted_mortality.xlsx
出力: data/static/age_adjusted_mortality_2020.json

対象: 6死因 × 47都道府県 × 男女
- 悪性新生物 (がん)
- 心疾患 (循環器)
- 脳血管疾患 (脳血管)
- 糖尿病 (糖尿病)
- 腎不全 (腎疾患)
- 肺炎 (呼吸器)

UIへの反映はしない (peer review 採択方針)。

Excel構造:
  シート 参考2(1): col 1=県名, col 2-25 = 全死因/悪性新生物/胃/大腸/肺の悪新/糖尿病 (各 男率/男順位/女率/女順位)
  シート 参考2(2): col 1=県名, col 2-19 = 心疾患/急性心筋梗塞/脳血管/脳梗塞/肺炎 (各 男率/男順位/女率/女順位)
  シート 参考2(3): col 1=県名, col 2-21 = 肝/腎不全/老衰/不慮事故/自殺 (各 男率/男順位/女率/女順位)

各死因は 4列ブロック (男率/男順位/女率/女順位)。
ヘッダ: row 3=大分類, row 4=男女, row 5=率/順位
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_age_adjusted_mortality' / 'r2_age_adjusted_mortality.xlsx'
OUT = ROOT / 'data' / 'static' / 'age_adjusted_mortality_2020.json'

# 抽出対象 6死因: (Excel列ヘッダ名, 出力キー名, シート名, 開始列)
# 列構造の再確認に基づき、ヘッダラベルを使って自動探索する
TARGET_CAUSES = {
    '悪性新生物＜腫瘍＞': '悪性新生物',
    '糖尿病': '糖尿病',
    '心疾患': '心疾患',
    '脳血管疾患': '脳血管疾患',
    '腎不全': '腎不全',
    '肺炎': '肺炎',
}


# 県名マッピング (xlsx ベースの短縮名 → 正式名)
PREF_MAP = {
    '北海道': '北海道',
    '青森': '青森県', '岩手': '岩手県', '宮城': '宮城県', '秋田': '秋田県',
    '山形': '山形県', '福島': '福島県',
    '茨城': '茨城県', '栃木': '栃木県', '群馬': '群馬県', '埼玉': '埼玉県',
    '千葉': '千葉県', '東京': '東京都', '神奈川': '神奈川県',
    '新潟': '新潟県', '富山': '富山県', '石川': '石川県', '福井': '福井県',
    '山梨': '山梨県', '長野': '長野県', '岐阜': '岐阜県', '静岡': '静岡県',
    '愛知': '愛知県', '三重': '三重県',
    '滋賀': '滋賀県', '京都': '京都府', '大阪': '大阪府', '兵庫': '兵庫県',
    '奈良': '奈良県', '和歌山': '和歌山県',
    '鳥取': '鳥取県', '島根': '島根県', '岡山': '岡山県', '広島': '広島県', '山口': '山口県',
    '徳島': '徳島県', '香川': '香川県', '愛媛': '愛媛県', '高知': '高知県',
    '福岡': '福岡県', '佐賀': '佐賀県', '長崎': '長崎県', '熊本': '熊本県',
    '大分': '大分県', '宮崎': '宮崎県', '鹿児島': '鹿児島県', '沖縄': '沖縄県',
    '全国': '全国',
}


def normalize_pref(s):
    """県名から空白文字を除去 + 正式名にマッピング"""
    if s is None: return None
    cleaned = str(s).strip().replace('\u3000', '').replace(' ', '')
    return PREF_MAP.get(cleaned, cleaned)


def find_cause_columns(ws):
    """row 3 のヘッダから死因の開始列を探索 → {大分類名: 開始列}"""
    columns = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if v and str(v).strip() in TARGET_CAUSES:
            columns[str(v).strip()] = c
    return columns


def extract_pref_rates(xlsx_path):
    """6死因 × 47県 × 男女 の率を抽出"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    # 結果格納: prefectures[pref][cause] = {male, female}
    pref_data = {}
    
    # 全国データも別途保存
    national = {}
    
    # 参考2(1)/(2)/(3) の3シート
    sheets = ['参考２（１）', '参考２（２）', '参考２（３）']
    
    for sn in sheets:
        ws = wb[sn]
        cause_cols = find_cause_columns(ws)
        if not cause_cols:
            print(f'  {sn}: 対象死因なし')
            continue
        print(f'  {sn}: 抽出対象 {list(cause_cols.keys())}')
        
        # データ行を走査 (row 6 = 全国, row 7+ = 47県)
        # 注記行 (注: ...) は無視
        for r in range(6, ws.max_row + 1):
            pref_raw = ws.cell(r, 1).value
            pref = normalize_pref(pref_raw)
            if not pref or '注' in pref or '）' in pref or pref.startswith('2'): 
                continue
            # 全国 or 47都道府県
            
            for cause_label, start_col in cause_cols.items():
                # 列レイアウト確認:
                #   start_col   = 男 率
                #   start_col+1 = 男 順位
                #   start_col+2 = 女 率
                #   start_col+3 = 女 順位
                # ただし全国の row 6 では順位が空だが値の位置は同じ
                male_rate = ws.cell(r, start_col).value
                male_rank = ws.cell(r, start_col + 1).value
                female_rate = ws.cell(r, start_col + 2).value
                female_rank = ws.cell(r, start_col + 3).value
                
                # 数値型チェック
                if not isinstance(male_rate, (int, float)) or not isinstance(female_rate, (int, float)):
                    continue
                
                out_cause = TARGET_CAUSES[cause_label]
                
                if pref == '全国':
                    national.setdefault(out_cause, {})['male'] = {'rate': male_rate, 'rank': None}
                    national[out_cause]['female'] = {'rate': female_rate, 'rank': None}
                else:
                    pref_data.setdefault(pref, {})[out_cause] = {
                        'male': {
                            'rate': male_rate,
                            'rank': male_rank if isinstance(male_rank, int) else None,
                        },
                        'female': {
                            'rate': female_rate,
                            'rank': female_rank if isinstance(female_rank, int) else None,
                        },
                        'total_simple_mean': round((male_rate + female_rate) / 2, 2),
                    }
    
    wb.close()
    return pref_data, national


def main():
    if not RAW.exists():
        print(f'ERROR: {RAW} not found.')
        print(f'Download: curl -sL "https://www.mhlw.go.jp/toukei/saikin/hw/jinkou/other/20sibou/xls/13.xlsx" -o {RAW}')
        return
    
    print(f'Phase 3-1 Option B: 年齢調整死亡率 2020 ETL 開始')
    print(f'入力: {RAW}')
    
    pref_data, national = extract_pref_rates(RAW)
    
    # 47県カバレッジ確認
    print(f'\n47県カバレッジ確認:')
    expected = 47
    found_prefs = sorted(pref_data.keys())
    print(f'  検出県数: {len(found_prefs)}/{expected}')
    
    # 6死因 × 47県の欠損確認
    causes_expected = list(TARGET_CAUSES.values())
    missing_count = 0
    for pref in found_prefs:
        for cause in causes_expected:
            if cause not in pref_data[pref]:
                missing_count += 1
                print(f'  ⚠️  欠損: {pref} - {cause}')
    print(f'  欠損総数: {missing_count} (期待: 0)')
    
    # 出力
    output = {
        'source': '令和5年度人口動態統計特殊報告 令和2年都道府県別年齢調整死亡率の概況',
        'source_url': 'https://www.mhlw.go.jp/toukei/saikin/hw/jinkou/other/20sibou/index.html',
        'year': 2020,
        'standard_population': '平成27年(2015年)モデル人口',
        'unit': '人口10万対',
        'method': '直接法による年齢調整死亡率',
        'causes': causes_expected,
        'notes': [
            '年齢調整死亡率は2020年(令和2年)時点',
            '既存の粗死亡率データ (vital_stats_pref.json, 2024年) とは時点が異なる',
            '基準人口は平成27年(2015年)モデル人口',
            '5年ごとに公表 (次回は令和7年(2025年)、2027年頃公表予定)',
            'total_simple_mean は男女の単純平均 (人口加重なし)。地域比較で用いる際は注意',
            'Bridge UIへの反映はPhase 3-1 Option Bでは未実施。別途設計判断後にUI実装予定',
            'rank は厚労省公表値 (該当死因における47県中の順位)',
        ],
        'extraction_date': '2026-04-29',
        'extraction_phase': 'Phase 3-1 Option B (ETL only, no UI integration)',
        'national': national,
        'prefectures': pref_data,
    }
    
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f'\n保存: {OUT}')
    print(f'  47県 × 6死因 × 男女別 = {sum(len(d) for d in pref_data.values())} エントリ')
    
    # ── Sanity check ──
    print(f'\n=== Sanity check ===')
    
    # 1. 沖縄糖尿病: 男 20.8 / 女 9.7
    oki_dm = pref_data.get('沖縄県', {}).get('糖尿病', {})
    print(f'\n1. 沖縄糖尿病:')
    print(f'   男 {oki_dm.get("male", {}).get("rate")} (期待: 20.8) — {"✅" if oki_dm.get("male", {}).get("rate") == 20.8 else "❌"}')
    print(f'   女 {oki_dm.get("female", {}).get("rate")} (期待: 9.7) — {"✅" if oki_dm.get("female", {}).get("rate") == 9.7 else "❌"}')
    print(f'   男順位 {oki_dm.get("male", {}).get("rank")} (期待: 2) — {"✅" if oki_dm.get("male", {}).get("rank") == 2 else "❌"}')
    print(f'   女順位 {oki_dm.get("female", {}).get("rank")} (期待: 2) — {"✅" if oki_dm.get("female", {}).get("rank") == 2 else "❌"}')
    
    # 2. 香川糖尿病: 男 21.2 (1位)
    kagawa_dm = pref_data.get('香川県', {}).get('糖尿病', {})
    print(f'\n2. 香川糖尿病 男:')
    print(f'   率 {kagawa_dm.get("male", {}).get("rate")} (期待: 21.2 / 47県最高) — {"✅" if kagawa_dm.get("male", {}).get("rate") == 21.2 else "❌"}')
    print(f'   順位 {kagawa_dm.get("male", {}).get("rank")} (期待: 1) — {"✅" if kagawa_dm.get("male", {}).get("rank") == 1 else "❌"}')
    
    # 3. 5県主要死因の値表示
    print(f'\n3. 5県 × 6死因 sanity:')
    target_prefs = ['沖縄県', '高知県', '北海道', '東京都', '大阪府']
    for p in target_prefs:
        d = pref_data.get(p, {})
        if not d: 
            print(f'   {p}: データなし')
            continue
        line = f'   {p:<6}: '
        for c in causes_expected:
            cd = d.get(c, {})
            m = cd.get('male', {}).get('rate', '-')
            f_ = cd.get('female', {}).get('rate', '-')
            line += f'{c[:4]}({m}/{f_}) '
        print(line)
    
    # 4. 列順序検証 (沖縄の値は予期する範囲内か)
    print(f'\n4. 列順序検証 (悪性新生物):')
    oki_can = pref_data.get('沖縄県', {}).get('悪性新生物', {})
    print(f'   沖縄 悪性新生物 男率: {oki_can.get("male", {}).get("rate")} (期待: 300-500の範囲)')
    print(f'   沖縄 悪性新生物 女率: {oki_can.get("female", {}).get("rate")} (期待: 150-250の範囲)')
    male_rate = oki_can.get('male', {}).get('rate', 0)
    female_rate = oki_can.get('female', {}).get('rate', 0)
    if 300 < male_rate < 500 and 150 < female_rate < 250:
        print(f'   ✅ 列順序正常 (男率 > 女率 = 期待)')
    else:
        print(f'   ❌ 列順序異常の可能性')
    
    print(f'\nETL 完了')


if __name__ == '__main__':
    main()
