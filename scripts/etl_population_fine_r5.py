#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""NDB発生率法(#44/45手術・#50/51在宅)用の細分年齢人口。
社人研は 0-4…85-89, 90-94, 95歳～ の20階級を持つが population_r5 は85-89で打切り
(90+/95+を欠落)。NDBは90+を1階級で持つため、両者を突合できる19階級
(0-4…85-89, 90+) で二次医療圏×年に集約する。muni→圏マップは etl_population_r5 を再利用。

出力: data/static/population_fine_r5.json  {areas:{code:{years:{Y:{bands19:[...]}}}}}
"""
import json
from pathlib import Path
from collections import defaultdict
import openpyxl

import etl_population_r5 as base

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / 'data' / 'static' / 'population_fine_r5.json'
# 19階級: 0-4…85-89 = col6-23, 90+ = col24(90-94)+col25(95+)
BAND_COLS = list(range(6, 24))  # 18列 = 0-4…85-89
COL_90_94, COL_95 = 24, 25


def add19(bands, row, w=1.0):
    for i, c in enumerate(BAND_COLS):
        bands[i] += base.num(row[c]) * w
    bands[18] += (base.num(row[COL_90_94]) + base.num(row[COL_95])) * w


def main():
    muni2hsa, prefix4, name2hsa = base.build_maps()
    areas = defaultdict(lambda: {y: [0.0] * 19 for y in base.YEARS})
    wb = openpyxl.load_workbook(base.POP, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] is None or row[4] is None:
            continue
        kubun = str(row[1]).strip() if row[1] is not None else ''
        if kubun in ('a', '1'):
            continue
        code = str(row[0]).strip().split('.')[0].zfill(5)
        pref = str(row[2]).strip() if row[2] else ''
        name = str(row[3]).strip() if row[3] else ''
        try:
            year = int(str(row[4]).replace('年', '').strip())
        except (ValueError, TypeError):
            continue
        if year not in base.YEARS:
            continue
        if code in base.SPECIAL_SPLIT:
            for hsa, w in base.SPECIAL_SPLIT[code]:
                add19(areas[hsa][year], row, w)
            continue
        hsa = base.resolve(code, pref, name, muni2hsa, prefix4, name2hsa)
        if not hsa:
            continue
        add19(areas[hsa][year], row)
    wb.close()

    out = {}
    for hsa, ys in areas.items():
        out[hsa] = {'years': {str(y): [round(x) for x in ys[y]] for y in base.YEARS}}
    payload = {
        'source': '国立社会保障・人口問題研究所 令和5年推計(90+分割・NDB突合用)',
        'note': '19階級 0-4…85-89,90+。90+=90-94+95歳～。発生率法(NDB)専用。',
        'years': base.YEARS, 'bands': 19, 'areaCount': len(out), 'areas': out,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
    print(f"[pop-fine] 圏={len(out)} out={OUT.name}")
    ya = out.get('2606', {}).get('years', {})
    if ya:
        b20 = ya['2020']
        print(f"[pop-fine] 山城南2020 総={sum(b20):.0f} 85-89={b20[17]:.0f} 90+={b20[18]:.0f}"
              f" (旧pop_r5 band17=85-89のみ)")


if __name__ == '__main__':
    main()
