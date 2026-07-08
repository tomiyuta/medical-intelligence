#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P3 ETL: 令和5年度 DPC退院患者調査 → 二次医療圏×医療機関×MDC別 退院患者数
医療需給総覧カルテ #25,26（MDC別退院患者数）/#68（MDC別退院患者数・医療機関シェア）をネイティブ再構築。

入力（厚労省 中医協DPC評価分科会 令和5年度退院患者調査）:
  MDC別医療機関別件数.xlsx (001468867) : 告示番号|通番|施設名|手術(無/有)|MDC01-18 のDPC患者数（件数シート）
  施設概要表.xlsx (001468611)          : 通番→市町村番号・都道府県・病院類型
市町村番号→二次医療圏は R6施設票（col8市区町村コード→col5二次医療圏名→area_master）から構築。

出力: data/static/dpc_mdc_r5.json
  areas[hsaCode] = {
    pref, area,
    totals: { total, dpcOnlyTotal, mdc: {code:{name,count,dpcOnly}} },
    facilities: [{name, ruikei, isDpc, total, mdc:{code:count}}]
  }
"""
import json, re
import openpyxl
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
DPC = Path('/Users/yutatomi/Downloads/01_投資・定量分析/MedicalCRM_Data/DPC退院患者調査_R5')
F_MDC = DPC / 'MDC別医療機関別件数.xlsx'
F_FAC = DPC / '施設概要表.xlsx'
F_DEMAND = DPC / '医療圏別MDC患者数.xlsx'   # 患者住所地(需要)ベース 二次医療圏×MDC
F_AMBUL = DPC / '救急搬送有無_医療機関別MDC別.xlsx'  # 救急車搬送有無 施設×MDC
F_LOS = DPC / '在院日数の状況.xlsx'   # 施設別×年度の退院患者数(件数)=真の総数・マスクなし
DB = ROOT / 'data' / 'medical_intelligence.db'  # dpc_enrichment_v2(平成30年度=2018)

# 秋田県: 医療需給総覧は3圏(県北/県央/県南)、DPCデータは旧8圏 → 集約マップ
AKITA_MAP = {
    '大館・鹿角': '県北', '北秋田': '県北', '能代・山本': '県北',
    '秋田周辺': '県央', '由利本荘・にかほ': '県央',
    '大仙・仙北': '県南', '横手': '県南', '湯沢・雄勝': '県南',
}
SHISETSU = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6' / 'R6_施設票.xlsx'
MASTER = ROOT / 'data' / 'hsa' / 'area_master.json'
OUT = ROOT / 'data' / 'static' / 'dpc_mdc_r5.json'

MDC = {
    '01': '神経系', '02': '眼科系', '03': '耳鼻咽喉科系', '04': '呼吸器系', '05': '循環器系',
    '06': '消化器系', '07': '筋骨格系', '08': '皮膚・皮下組織', '09': '乳房', '10': '内分泌・栄養・代謝',
    '11': '腎・尿路系', '12': '女性生殖器系', '13': '血液・造血器', '14': '新生児・先天性', '15': '小児',
    '16': '外傷・熱傷・中毒', '17': '精神', '18': 'その他',
}
MDC_KEYS = [f'{i:02d}' for i in range(1, 19)]


def to_int(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip().replace(',', '')
    if s in ('', '-', '－', '*'): return 0
    try: return int(float(s))
    except ValueError: return 0


def build_muni2hsa():
    master = {(m['pref_code'], m['siteArea']): m['code']
              for m in json.load(open(MASTER, encoding='utf-8'))['areas'] if m.get('siteArea')}
    wb = openpyxl.load_workbook(SHISETSU, read_only=True, data_only=True)
    ws = wb['Sheet1']
    muni2hsa = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        mc, pc, an = row[8], row[3], row[5]
        if mc and pc and an:
            mc = str(mc).strip().split('.')[0].zfill(5)
            pc = str(pc).strip().split('.')[0].zfill(2)
            hsa = master.get((pc, str(an).strip()))
            if hsa:
                muni2hsa[mc] = hsa
    wb.close()
    return muni2hsa


def norm_name(s):
    """施設名の正規化（法人格・記号・空白除去）で名寄せ。"""
    s = str(s)
    s = re.sub(r'(医療法人社団|医療法人財団|医療法人|社会福祉法人|公益財団法人|公益社団法人|一般財団法人|'
              r'一般社団法人|独立行政法人|国立研究開発法人|地方独立行政法人|社会医療法人|特定医療法人|'
              r'学校法人|国家公務員共済組合連合会|地方公務員共済組合|日本赤十字社|厚生農業協同組合連合会|'
              r'全国厚生農業協同組合連合会|労働者健康安全機構)', '', s)
    s = re.sub(r'[\s　・（）\(\)]', '', s)
    return s.strip()


def build_name2hsa():
    """R6施設票の 施設名(正規化) → hsaコード。市町村番号の年次差(政令市区再編等)を吸収するフォールバック。"""
    master = {(m['pref_code'], m['siteArea']): m['code']
              for m in json.load(open(MASTER, encoding='utf-8'))['areas'] if m.get('siteArea')}
    wb = openpyxl.load_workbook(SHISETSU, read_only=True, data_only=True)
    ws = wb['Sheet1']
    name2hsa = {}
    for r in ws.iter_rows(min_row=7, values_only=True):
        nm, pc, an = r[2], r[3], r[5]
        if nm and pc and an:
            pc = str(pc).strip().split('.')[0].zfill(2)
            h = master.get((pc, str(an).strip()))
            if h:
                name2hsa[norm_name(nm)] = h
    wb.close()
    return name2hsa


def load_facility_master():
    """通番 → {muni, ruikei, isDpc}"""
    wb = openpyxl.load_workbook(F_FAC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    fac = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        tsuban = row[1]
        if tsuban is None:
            continue
        tsuban = str(tsuban).strip().split('.')[0]
        muni = str(row[2]).strip().split('.')[0].zfill(5) if row[2] else None
        ruikei = str(row[5]).strip() if row[5] else ''
        name = str(row[4]).strip() if row[4] else ''
        fac[tsuban] = {'muni': muni, 'ruikei': ruikei, 'isDpc': 'DPC' in ruikei, 'name': name}
    wb.close()
    return fac


def norm_area(s):
    s = str(s).strip()
    s = re.sub(r'(地域|医療圏|区域|圏)$', '', s)
    s = re.sub(r'[・･\s　]', '', s)
    return s


def load_demand():
    """患者住所地(需要)ベース 二次医療圏×MDC を hsaコードへ集約。名前正規化＋秋田8→3集約。"""
    master = json.load(open(MASTER, encoding='utf-8'))['areas']
    pref_full = {}
    for m in master:
        p = m['pref']
        short = p[:-1] if p.endswith(('県', '都', '府')) else p
        pref_full[short] = p
    pref_full['北海道'] = '北海道'
    by_pref = defaultdict(dict)   # pref -> normArea -> code
    name_by_pref = defaultdict(dict)  # pref -> exactName -> code
    for m in master:
        by_pref[m['pref']][norm_area(m['area'])] = m['code']
        name_by_pref[m['pref']][m['area']] = m['code']

    wb = openpyxl.load_workbook(F_DEMAND, read_only=True, data_only=True)
    ws = wb['二次医療圏_MDC別']
    demand = defaultdict(lambda: {'total': 0, 'mdc': defaultdict(int)})
    unmatched = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        pref = str(row[0]).strip() if row[0] else None
        area = str(row[1]).strip() if row[1] else None
        if not pref or not area or area == '不明':
            continue
        pf = pref_full.get(pref, pref if pref.endswith(('県', '都', '府', '道')) else pref + '県')
        # 秋田は8→3集約
        if pf == '秋田県' and area in AKITA_MAP:
            code = name_by_pref[pf].get(AKITA_MAP[area])
        else:
            code = by_pref[pf].get(norm_area(area))
        if not code:
            unmatched.append((pref, area))
            continue
        for j, k in enumerate(MDC_KEYS):
            v = to_int(row[2 + j])
            demand[code]['mdc'][k] += v
            demand[code]['total'] += v
    wb.close()
    if unmatched:
        print(f"[etl] 需要未突合 {len(unmatched)}: {unmatched[:8]}", flush=True)
    return demand


def load_trend():
    """通番 → {年度: 退院患者数}。在院日数ファイル(2019-2023件数)＋dpc_enrichment_v2(2018)。真の総数・マスクなし。"""
    trend = defaultdict(dict)
    # 在院日数: 令和元(col3)=2019 … 令和5(col35)=2023, 各年度8列間隔
    year_cols = {2019: 3, 2020: 11, 2021: 19, 2022: 27, 2023: 35}
    wb = openpyxl.load_workbook(F_LOS, read_only=True, data_only=True)
    ws = wb['在院日数の状況']
    for row in ws.iter_rows(min_row=4, values_only=True):
        tsuban = row[1]
        if tsuban is None or not str(tsuban).strip():
            continue
        t = str(tsuban).strip().split('.')[0]
        for y, c in year_cols.items():
            v = to_int(row[c] if len(row) > c else 0)
            if v > 0:
                trend[t][y] = v
    wb.close()
    # 2018(平成30年度) は dpc_enrichment_v2 (dpc_id=通番)
    import sqlite3
    con = sqlite3.connect(DB)
    for dpc_id, v in con.execute("SELECT dpc_id, 平成30年度 FROM dpc_enrichment_v2 WHERE 平成30年度 IS NOT NULL"):
        t = str(dpc_id).strip()
        if v:
            trend[t][2018] = int(v)
    con.close()
    return trend


def resolve_hsa(tsuban, facmaster, muni2hsa, name2hsa, name_hint=''):
    """通番→hsa。市町村番号(muni2hsa)優先、失敗時は施設名(name2hsa)でフォールバック。"""
    meta = facmaster.get(tsuban)
    if meta and meta['muni']:
        h = muni2hsa.get(meta['muni'])
        if h:
            return h
    nm = name_hint or (meta['name'] if meta else '')
    return name2hsa.get(norm_name(nm)) if nm else None


def main():
    muni2hsa = build_muni2hsa()
    name2hsa = build_name2hsa()
    facmaster = load_facility_master()
    demand = load_demand()
    trend = load_trend()
    print(f"[etl] muni2hsa={len(muni2hsa)} 施設概要={len(facmaster)} 需要圏={len(demand)}", flush=True)

    # 施設別MDC別（手術無/有を合算。手術有り件数も別途集計）
    wb = openpyxl.load_workbook(F_MDC, read_only=True, data_only=True)
    ws = wb['件数']
    fac_mdc = defaultdict(lambda: {'name': '', 'mdc': defaultdict(int), 'surgery': 0})
    cur = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        tsuban = row[1]
        if tsuban is not None and str(tsuban).strip():
            cur = str(tsuban).strip().split('.')[0]
            fac_mdc[cur]['name'] = str(row[2]).strip() if row[2] else ''
        if cur is None:
            continue
        rowsum = 0
        for j, k in enumerate(MDC_KEYS):
            v = to_int(row[4 + j])
            fac_mdc[cur]['mdc'][k] += v
            rowsum += v
        if str(row[3]).strip() == '有り':          # 手術有り行 → 手術件数
            fac_mdc[cur]['surgery'] += rowsum
    wb.close()
    print(f"[etl] DPC施設(MDC別)={len(fac_mdc)}", flush=True)

    # 救急車搬送 施設別（救急搬送列=col3,5,7,… の合計）
    amb = defaultdict(int)
    wb = openpyxl.load_workbook(F_AMBUL, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=4, values_only=True):
        tsuban = row[1]
        if tsuban is None or not str(tsuban).strip():
            continue
        t = str(tsuban).strip().split('.')[0]
        for c in range(3, 39, 2):   # 救急車搬送列（MDC01-18）
            amb[t] += to_int(row[c] if len(row) > c else 0)
    wb.close()
    print(f"[etl] 救急搬送施設={len(amb)}", flush=True)

    areas = defaultdict(lambda: {'facilities': [],
                                 'mdc': defaultdict(int), 'mdcDpc': defaultdict(int),
                                 'total': 0, 'dpcOnlyTotal': 0})
    unmatched = 0
    for tsuban, fm in fac_mdc.items():
        meta = facmaster.get(tsuban)
        hsa = resolve_hsa(tsuban, facmaster, muni2hsa, name2hsa, fm.get('name', ''))
        if not meta or not hsa:
            unmatched += 1
            continue
        total = sum(fm['mdc'].values())
        if total == 0:
            continue
        is_dpc = meta['isDpc']
        a = areas[hsa]
        a['facilities'].append({
            'name': fm['name'], 'ruikei': meta['ruikei'], 'isDpc': is_dpc,
            'total': total, 'mdc': {k: fm['mdc'][k] for k in MDC_KEYS if fm['mdc'][k] > 0},
            'surgery': fm.get('surgery', 0), 'ambulance': amb.get(tsuban, 0),
        })
        a['total'] += total
        if is_dpc:
            a['dpcOnlyTotal'] += total
        for k in MDC_KEYS:
            a['mdc'][k] += fm['mdc'][k]
            if is_dpc:
                a['mdcDpc'][k] += fm['mdc'][k]

    # 退院患者数の年度推移を圏へ集約（DPC対象病院のみ=PDF#25と同スコープ・通番→市町村→圏）
    hsa_trend = defaultdict(lambda: defaultdict(int))
    YEARS = [2018, 2019, 2020, 2021, 2022, 2023]
    for tsuban, ymap in trend.items():
        meta = facmaster.get(tsuban)
        if not meta or not meta['isDpc']:   # DPC対象病院のみ
            continue
        hsa = resolve_hsa(tsuban, facmaster, muni2hsa, name2hsa)
        if not hsa:
            continue
        for y, v in ymap.items():
            hsa_trend[hsa][y] += v

    pref_name = {m['code'][:2]: m['pref'] for m in json.load(open(MASTER, encoding='utf-8'))['areas']}
    area_name = {m['code']: m['area'] for m in json.load(open(MASTER, encoding='utf-8'))['areas']}
    out = {}
    for hsa, a in areas.items():
        a['facilities'].sort(key=lambda x: -x['total'])
        dem = demand.get(hsa)
        flow = None
        if dem and dem['total'] > 0:
            # 供給(所在地)=住所地不問で当該圏医療機関のDPC患者数=需要ファイルの所在地に相当する住所地対比。
            # 需要(住所地)=当該圏住民のDPC患者数(clean)。完結率=供給/需要(≒PDF定義 所在地/住所地)。
            flow = {
                'demand': dem['total'],
                'demandMdc': {k: dem['mdc'][k] for k in MDC_KEYS if dem['mdc'][k] > 0},
            }
        tr = hsa_trend.get(hsa)
        trend_out = [{'year': y, 'count': tr[y]} for y in YEARS if tr and tr.get(y)] if tr else []
        out[hsa] = {
            'pref': pref_name.get(hsa[:2], ''), 'area': area_name.get(hsa, ''),
            'totals': {
                'total': a['total'], 'dpcOnlyTotal': a['dpcOnlyTotal'],
                'mdc': {k: {'name': MDC[k], 'count': a['mdc'][k], 'dpcOnly': a['mdcDpc'][k]}
                        for k in MDC_KEYS if a['mdc'][k] > 0},
            },
            'facilities': a['facilities'],
            'flow': flow,
            'trend': trend_out,   # DPC対象病院の退院患者数 年度推移(2018-2023, 真の総数)
        }

    payload = {
        'source': '厚労省 令和5年度DPC導入の影響評価に係る調査「退院患者調査」（2023年度）',
        'note': 'MDC別退院患者数（DPC対象・準備・出来高算定病院）。施設×手術無/有を合算。DPC病院のみ集計も併載。市町村→二次医療圏はR6施設票由来。',
        'mdcLabels': MDC,
        'areaCount': len(out),
        'areas': out,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
    print(f"[etl] areas={len(out)} 未突合施設={unmatched} out={OUT.stat().st_size/1e6:.1f}MB")


if __name__ == '__main__':
    main()
