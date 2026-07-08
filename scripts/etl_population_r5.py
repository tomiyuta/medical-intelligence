#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P4 ETL(基盤): 社人研 年齢5歳階級別 将来推計人口(令和5年推計) → 二次医療圏別に集約
医療需給総覧 #28(人口推計)/#29(高齢化)、受療率法推計(#30-)の人口(分母)基礎データ。

市町村→二次医療圏マップ(完全)= コード(R6施設票 muni→圏 + 政令市の先頭4桁)＋名前(area_demographics 圏→市町村名)。
社人研 列: col0=市区町村コード, col1=市などの別(a=都道府県計,除外), col2=都道府県, col3=市区町村,
  col4=年(「2020年」形式), col5=総数, col6-23=5歳階級18(0-4…85+), col73=0-14, col74=15-64, col75=65+

出力: data/static/population_r5.json  { areas: {hsaCode:{pref,area, years:{Y:{total,a0_14,a15_64,a65,a75, bands:[18]}}}} }
"""
import openpyxl, json
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
POP = Path('/Users/yutatomi/Downloads/01_投資・定量分析/MedicalCRM_Data/P4_将来推計/社人研_年齢別将来人口.xlsx')
SHISETSU = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6' / 'R6_施設票.xlsx'
MASTER = ROOT / 'data' / 'hsa' / 'area_master.json'
AREA_DEMO = ROOT / 'data' / 'static' / 'area_demographics.json'
OUT = ROOT / 'data' / 'static' / 'population_r5.json'

YEARS = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
AGE_COLS = list(range(6, 24))   # 0-4 … 85+ の18階級
C_TOTAL, C_0_14, C_15_64, C_65 = 5, 73, 74, 75
# 75+ = 75-79,80-84,85+ = col21,22,23
C_75 = [21, 22, 23]
# 福島 社人研の集約地域(浜通り=いわき+相双を1集約)。国調(カルテ#4)人口比で按分:
#   相双 119,577 / いわき 332,931 / 計452,508
SPECIAL_SPLIT = {'07999': [('0706', 119577 / 452508), ('0707', 332931 / 452508)]}


def num(v):
    return v if isinstance(v, (int, float)) else 0


def build_maps():
    master = json.load(open(MASTER, encoding='utf-8'))['areas']
    key2hsa = {(m['pref_code'], m['siteArea']): m['code'] for m in master if m.get('siteArea')}
    prefarea2hsa = {(m['pref'], m['area']): m['code'] for m in master}
    prefarea2hsa.update({(m['pref'], m['siteArea']): m['code'] for m in master if m.get('siteArea')})

    # コードマップ(R6施設票)
    muni2hsa, prefix4 = {}, {}
    wb = openpyxl.load_workbook(SHISETSU, read_only=True, data_only=True)
    ws = wb['Sheet1']
    for r in ws.iter_rows(min_row=7, values_only=True):
        mc, pc, an = r[8], r[3], r[5]
        if mc and pc and an:
            mc = str(mc).strip().split('.')[0].zfill(5)
            pc = str(pc).strip().split('.')[0].zfill(2)
            h = key2hsa.get((pc, str(an).strip()))
            if h:
                muni2hsa[mc] = h
                prefix4.setdefault(mc[:4], h)
    wb.close()

    # 名前マップ(area_demographics 圏→市町村名)
    name2hsa = {}
    for a in json.load(open(AREA_DEMO, encoding='utf-8')):
        h = prefarea2hsa.get((a['pref'], a['area']))
        if not h:
            continue
        for mu in a.get('munis', []):
            name2hsa[(a['pref'], mu['name'])] = h
    return muni2hsa, prefix4, name2hsa


def resolve(code, pref, name, muni2hsa, prefix4, name2hsa):
    return (muni2hsa.get(code) or name2hsa.get((pref, name)) or prefix4.get(code[:4]))


def add_to(a, row, weight=1.0):
    a['total'] += num(row[C_TOTAL]) * weight
    a['a0_14'] += num(row[C_0_14]) * weight
    a['a15_64'] += num(row[C_15_64]) * weight
    a['a65'] += num(row[C_65]) * weight
    a['a75'] += sum(num(row[c]) for c in C_75) * weight
    for i, c in enumerate(AGE_COLS):
        a['bands'][i] += num(row[c]) * weight


def main():
    muni2hsa, prefix4, name2hsa = build_maps()
    print(f"[pop] コードマップ={len(muni2hsa)} 政令市prefix={len(prefix4)} 名前マップ={len(name2hsa)}", flush=True)

    areas = defaultdict(lambda: {y: {'total': 0, 'a0_14': 0, 'a15_64': 0, 'a65': 0, 'a75': 0,
                                     'bands': [0] * 18} for y in YEARS})
    unc_pop, cov_pop = 0.0, 0.0
    unc = {}
    wb = openpyxl.load_workbook(POP, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] is None or row[4] is None:
            continue
        kubun = str(row[1]).strip() if row[1] is not None else ''
        if kubun in ('a', '1'):   # 都道府県計・政令市の市集約(区で計上するため二重計上回避)
            continue
        code = str(row[0]).strip().split('.')[0].zfill(5)
        pref = str(row[2]).strip() if row[2] else ''
        name = str(row[3]).strip() if row[3] else ''
        try:
            year = int(str(row[4]).replace('年', '').strip())
        except (ValueError, TypeError):
            continue
        if year not in YEARS:
            continue
        tot = num(row[C_TOTAL])
        # 福島浜通り集約は複数圏へ按分
        if code in SPECIAL_SPLIT:
            for hsa, w in SPECIAL_SPLIT[code]:
                add_to(areas[hsa][year], row, w)
            if year == 2020:
                cov_pop += tot
            continue
        hsa = resolve(code, pref, name, muni2hsa, prefix4, name2hsa)
        if not hsa:
            if year == 2020:
                unc_pop += tot; unc[code] = (pref, name, tot)
            continue
        if year == 2020:
            cov_pop += tot
        add_to(areas[hsa][year], row)
    wb.close()

    out = {}
    master = {m['code']: m for m in json.load(open(MASTER, encoding='utf-8'))['areas']}
    for hsa, ys in areas.items():
        m = master.get(hsa, {})
        out[hsa] = {
            'pref': m.get('pref', ''), 'area': m.get('area', ''),
            'years': {str(y): {k: (round(v) if not isinstance(v, list) else [round(x) for x in v])
                               for k, v in ys[y].items()} for y in YEARS},
        }
    payload = {
        'source': '国立社会保障・人口問題研究所「日本の地域別将来推計人口 令和5(2023)年推計」',
        'note': '市区町村別5歳階級推計を二次医療圏へ集約。2020年は推計基準人口。',
        'years': YEARS, 'areaCount': len(out), 'areas': out,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
    print(f"[pop] 圏={len(out)} 被覆人口={cov_pop:,.0f} 未被覆={unc_pop:,.0f}"
          f" ({unc_pop/(cov_pop+unc_pop)*100:.2f}%)")
    if unc:
        print("[pop] 未被覆:", sorted(unc.items(), key=lambda x: -x[1][2])[:6])
    ya = out.get('2606', {}).get('years', {}).get('2020', {})
    print(f"[pop] 検証 山城南2020: 総人口={ya.get('total')} (PDF#4国調=121,118)")


if __name__ == '__main__':
    main()
