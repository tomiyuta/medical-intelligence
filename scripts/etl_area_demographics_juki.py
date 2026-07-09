#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETL: 住民基本台帳(総務省) → data/static/area_demographics.json 全面再生成

これまで area_demographics.json は外部生成(fe0d01c「Pottech式」)でETLがリポジトリに
無く、政令指定都市20市が丸ごと欠落していた(munis合算 96.9M vs 実際 124.3M)。
本スクリプトは住基の一次資料から全市区町村を再生成し、以下を根治する:
  1. 政令指定都市20市の追加(川崎市のみ圏が区で分かれるため区単位、他19市は市単位)
  2. 泊村(北海道後志)が北方領土の国後郡泊村(人口0)と名前衝突で全フィールド0だったバグ
  3. 上郡町(兵庫)の値が住基R7に存在しない古い値だったバグ

入力(e-Stat 住民基本台帳に基づく人口、人口動態及び世帯数調査 令和7年 toukei=00200241):
  data/raw/source/13_住民基本台帳/R7_2503_市区町村別人口動態世帯.xlsx
    表25-03【総計】市区町村別人口、人口動態及び世帯数(2025-01-01現在、動態は令和6年中)
    https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040306653&fileKind=0
  data/raw/source/13_住民基本台帳/R7_2504_市区町村別年齢階級別人口.xlsx
    表25-04【総計】市区町村別年齢階級別人口(2025-01-01現在、5歳階級21区分)
    https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040306654&fileKind=0
  data/static/area_demographics.json … 既存の(pref,市町村名)→二次医療圏マップ源
  data/hsa/area_master.json          … 圏コード→(pref,area)名
  data/static/age_pyramid.json       … 検証(同じ住基総計由来の県・全国人口)

市町村→圏のマッピング:
  - 既存 area_demographics の (pref, name) を継承(320圏・1,721市町村)
  - 政令指定都市は R6病床機能報告 施設票の区コード→圏名で検証済みの固定表(SEIREI_HSA)
  - 北方領土6村(人口0・所属圏なし)は除外
  - それ以外で未マッピングの行があれば異常終了(silent fallback禁止)

検証: 全国munis合算 ≈ age_pyramid national(±0.5%)、各県合算 ≈ age_pyramid(±0.5%)、
      munis合算 = 住基ファイル全国行(北方領土0人を除き完全一致)

出力: data/static/area_demographics.json
  [{pref, area, munis:[{name,pop,p15,p65,births,deaths,nc,hh,aging}]}]
  トップレベルは(pref,area)文字列ソート、munisはpop降順。compact(改行なし)。
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw' / 'source' / '13_住民基本台帳'
DOTAI = RAW / 'R7_2503_市区町村別人口動態世帯.xlsx'
NENREI = RAW / 'R7_2504_市区町村別年齢階級別人口.xlsx'
MASTER = ROOT / 'data' / 'hsa' / 'area_master.json'
OUT = ROOT / 'data' / 'static' / 'area_demographics.json'
PYRAMID = ROOT / 'data' / 'static' / 'age_pyramid.json'

# 政令指定都市 → 二次医療圏コード(R6施設票の区コード→圏で全区一致を確認済み 2026-07-09)
# 川崎市のみ区で圏が分かれる(南部=川崎・幸・中原/北部=高津・多摩・宮前・麻生)ため区単位。
SEIREI_HSA = {
    '01100': '0104',  # 札幌市 → 札幌
    '04100': '0403',  # 仙台市 → 仙台
    '11100': '1104',  # さいたま市 → さいたま
    '12100': '1201',  # 千葉市 → 千葉
    '14100': '1412',  # 横浜市 → 横浜
    '14150': '1410',  # 相模原市 → 相模原
    '15100': '1502',  # 新潟市 → 新潟
    '22100': '2205',  # 静岡市 → 静岡
    '22130': '2208',  # 浜松市 → 西部
    '23100': '2313',  # 名古屋市 → 名古屋・尾張中部
    '26100': '2604',  # 京都市 → 京都・乙訓
    '27100': '2708',  # 大阪市 → 大阪市
    '27140': '2706',  # 堺市 → 堺市
    '28100': '2801',  # 神戸市 → 神戸
    '33100': '3301',  # 岡山市 → 県南東部
    '34100': '3401',  # 広島市 → 広島
    '40100': '4012',  # 北九州市 → 北九州
    '40130': '4001',  # 福岡市 → 福岡・糸島
    '43100': '4312',  # 熊本市 → 熊本・上益城
    '14131': '1405',  # 川崎市川崎区 → 川崎南部
    '14132': '1405',  # 川崎市幸区 → 川崎南部
    '14133': '1405',  # 川崎市中原区 → 川崎南部
    '14134': '1404',  # 川崎市高津区 → 川崎北部
    '14135': '1404',  # 川崎市多摩区 → 川崎北部
    '14136': '1404',  # 川崎市宮前区 → 川崎北部
    '14137': '1404',  # 川崎市麻生区 → 川崎北部
}
SEIREI_CITIES = ['札幌市', '仙台市', 'さいたま市', '千葉市', '横浜市', '川崎市', '相模原市',
                 '新潟市', '静岡市', '浜松市', '名古屋市', '京都市', '大阪市', '堺市', '神戸市',
                 '岡山市', '広島市', '北九州市', '福岡市', '熊本市']
# 北方領土6村(住基人口0・二次医療圏なし)
HOPPO = {'01695', '01696', '01697', '01698', '01699', '01700'}
# 住基ファイルの異体字 → area_demographics の表記
KYUJI = {'須惠町': '須恵町'}


def ival(v):
    return int(v) if isinstance(v, (int, float)) else 0


def strip_gun(name):
    """町村行の「郡名」プレフィックスを剥がす(市・区は郡を持たない: 大和郡山市等を誤爆しない)"""
    if name.endswith('市') or name.endswith('区'):
        return name
    m = re.match(r'^.+?郡(.+)$', name)
    name = m.group(1) if m else name
    return KYUJI.get(name, name)


def load_juki():
    """住基2表 → {code5: {pref,name,pop,hh,births,deaths,p15,p65}}, 全国行"""
    munis, national = {}, None
    wb = openpyxl.load_workbook(DOTAI, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=7, values_only=True):
        pref = str(r[1]).strip() if r[1] else ''
        name = str(r[2]).strip() if r[2] else ''
        rec = {'pop': ival(r[5]), 'hh': ival(r[6]), 'births': ival(r[10]), 'deaths': ival(r[16])}
        if pref == '合計':
            national = rec
            continue
        if not r[0] or name in ('', '-', '島しょ') or name.endswith('郡'):
            continue  # 県計・郡計・東京島嶼部の集計行(構成町村は別行で計上)
        code = str(r[0]).strip()[:5]
        munis[code] = {'pref': pref, 'name': name, **rec}
    wb.close()

    wb = openpyxl.load_workbook(NENREI, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[0] or str(r[3]).strip() != '計':
            continue
        code = str(r[0]).strip()[:5]
        if code in munis:
            bands = [ival(r[c]) for c in range(5, 26)]  # 0-4 … 100+ の21階級
            munis[code]['p15'] = sum(bands[:3])
            munis[code]['p65'] = sum(bands[13:])
    wb.close()

    missing_age = [c for c, d in munis.items() if 'p65' not in d]
    if missing_age:
        sys.exit(f"[juki] 年齢階級ファイルに無いコード: {missing_age[:10]}")
    return munis, national


def main():
    munis, national = load_juki()
    print(f"[juki] 市区町村行={len(munis)} 全国={national['pop']:,}")

    master = {m['code']: m for m in json.load(open(MASTER, encoding='utf-8'))['areas']}
    prev = json.load(open(OUT, encoding='utf-8'))
    name2area = {(a['pref'], m['name']): (a['pref'], a['area'])
                 for a in prev for m in a['munis']}
    prev_vals = {(a['pref'], m['name']): m for a in prev for m in a['munis']}

    # 市区町村行 → 圏へ振り分け
    is_ward = lambda n: any(n.startswith(c) and n != c for c in SEIREI_CITIES)
    grouped, unmapped = {}, []
    for code, d in munis.items():
        if code in HOPPO:
            continue
        if code in SEIREI_HSA:
            m = master.get(SEIREI_HSA[code])
            if not m:
                sys.exit(f"[juki] area_masterに圏コード{SEIREI_HSA[code]}が無い({d['name']})")
            key = (m['pref'], m['area'])
        elif is_ward(d['name']) or d['name'] in SEIREI_CITIES:
            continue  # 政令市の区(川崎以外)と川崎市の市計行 → SEIREI_HSA側で計上済み
        else:
            key = name2area.get((d['pref'], strip_gun(d['name'])))
            if not key:
                unmapped.append((code, d['pref'], d['name'], d['pop']))
                continue
        grouped.setdefault(key, []).append(d)
    if unmapped:
        for u in unmapped:
            print(f"[juki] 未マッピング: {u}")
        sys.exit(f"[juki] {len(unmapped)}行が二次医療圏に紐付かない(全行の紐付けが必須)")

    out = []
    for (pref, area), ds in sorted(grouped.items()):
        rows = []
        for d in sorted(ds, key=lambda x: -x['pop']):
            name = strip_gun(d['name']) if not is_ward(d['name']) else d['name']
            rows.append({'name': name, 'pop': d['pop'], 'p15': d['p15'], 'p65': d['p65'],
                         'births': d['births'], 'deaths': d['deaths'],
                         'nc': d['births'] - d['deaths'], 'hh': d['hh'],
                         'aging': round(d['p65'] / d['pop'] * 100, 1) if d['pop'] else 0})
        out.append({'pref': pref, 'area': area, 'munis': rows})

    # ---- 検証 ----
    tot = sum(m['pop'] for a in out for m in a['munis'])
    assert tot == national['pop'], f"munis合算{tot:,} ≠ 住基全国{national['pop']:,}"
    ap = json.load(open(PYRAMID, encoding='utf-8'))
    ap_nat = sum(ap['national']['male']) + sum(ap['national']['female'])
    dev = (tot - ap_nat) / ap_nat * 100
    print(f"[検証] 全国munis合算={tot:,} vs age_pyramid={ap_nat:,} ({dev:+.4f}%)")
    assert abs(dev) < 0.5
    bad = []
    for pref in sorted({a['pref'] for a in out}):
        s = sum(m['pop'] for a in out if a['pref'] == pref for m in a['munis'])
        p = ap['prefectures'].get(pref)
        apv = sum(p['male']) + sum(p['female']) if p else 0
        d = (s - apv) / apv * 100 if apv else 999
        if abs(d) >= 0.5:
            bad.append((pref, s, apv, d))
    if bad:
        for b in bad:
            print(f"[検証] 県合算乖離: {b[0]} munis={b[1]:,} pyramid={b[2]:,} ({b[3]:+.3f}%)")
        sys.exit('[検証] 県合算が±0.5%を超過')
    kyoto = sum(m['pop'] for a in out if a['pref'] == '京都府' for m in a['munis'])
    print(f"[検証] 47県すべて±0.5%以内 / 京都府={kyoto:,} (基準2,471,929)")

    # ---- 差分レポート ----
    new_areas = {(a['pref'], a['area']) for a in out} - {(a['pref'], a['area']) for a in prev}
    cur_vals = {(a['pref'], m['name']): m for a in out for m in a['munis']}
    added = sorted(set(cur_vals) - set(prev_vals))
    removed = sorted(set(prev_vals) - set(cur_vals))
    changed = sorted(k for k in set(cur_vals) & set(prev_vals) if cur_vals[k] != prev_vals[k])
    print(f"[差分] 新規圏={len(new_areas)} 追加muni={len(added)} 変更={len(changed)} 削除={len(removed)}")
    for k in added:
        print(f"  + {k[0]} {k[1]} pop={cur_vals[k]['pop']:,}")
    for k in changed:
        print(f"  ~ {k[0]} {k[1]} {prev_vals[k]['pop']:,}→{cur_vals[k]['pop']:,}")
    for k in removed:
        print(f"  - {k[0]} {k[1]}")

    OUT.write_text(json.dumps(out, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    print(f"[out] {OUT} 圏={len(out)} muni={len(cur_vals)}")


if __name__ == '__main__':
    main()
