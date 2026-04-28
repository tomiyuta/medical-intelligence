#!/usr/bin/env python3
"""
etl_ndb_checkup_bins_v2.py — NDB第10回 特定健診 検査値階層別分布 ETL v2

v1 (etl_ndb_checkup_bins.py) との違い:
- v1: 男女合算・全年齢合算でのみ保持 (年齢標準化不可)
- v2: 性 (male/female) × 年齢階級 (40-44, 45-49, ..., 70-74) を完全保持
- 標準人口テーブルも同時生成 (Phase 2C-1)

入力: data/raw_ndb_checkup/{metric}_pref.xlsx
出力:
- data/static/ndb_checkup_bins_v2.json (階級分布、性年齢別保持)
- data/static/ndb_checkup_standard_pop.json (NDB内標準人口)
- data/static/ndb_checkup_risk_rates_standardized.json (粗率 + 標準化率 + delta)

Excel構造:
  row 4: ヘッダー (40～44, 45～49, ..., 70～74, 中計, 40～44(女), ..., 中計(女))
  col 1: 都道府県 (continued)
  col 2: 検査値階層
  col 3-9: 男 7階級
  col 10: 男 中計
  col 11-17: 女 7階級
  col 18: 女 中計
"""
import json
import openpyxl
import collections
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_ndb_checkup'
OUT_BINS_V2 = ROOT / 'data' / 'static' / 'ndb_checkup_bins_v2.json'
OUT_STD_POP = ROOT / 'data' / 'static' / 'ndb_checkup_standard_pop.json'
OUT_RATES_STD = ROOT / 'data' / 'static' / 'ndb_checkup_risk_rates_standardized.json'

FILE_METRIC = {
    'BMI_pref.xlsx': 'BMI',
    'HbA1c_pref.xlsx': 'HbA1c',
    'SBP_pref.xlsx': '収縮期血圧',
    'LDL_pref.xlsx': 'LDL',
    'UrineProtein_pref.xlsx': '尿蛋白',
}

# 年齢階級 (col index → label)
AGE_GROUPS = ['40-44', '45-49', '50-54', '55-59', '60-64', '65-69', '70-74']

RISK_DEFS = {
    'BMI': {
        'risk_key': 'bmi_ge_25',
        'risk_label': 'BMI ≥25 (肥満)',
        'risk_bins': ['25.0以上30.0未満', '30.0以上35.0未満', '35.0以上40.0未満', '40.0以上'],
        'unit': 'kg/m²',
    },
    'HbA1c': {
        'risk_key': 'hba1c_ge_6_5',
        'risk_label': 'HbA1c ≥6.5% (糖尿病型)',
        'risk_bins': ['6.5以上8.0未満', '8.0以上8.4未満', '8.4以上'],
        'unit': '%',
    },
    '収縮期血圧': {
        'risk_key': 'sbp_ge_140',
        'risk_label': '収縮期血圧 ≥140 mmHg (高血圧)',
        'risk_bins': ['140以上160未満', '160以上180未満', '180以上'],
        'unit': 'mmHg',
    },
    'LDL': {
        'risk_key': 'ldl_ge_140',
        'risk_label': 'LDL ≥140 mg/dL (脂質異常症)',
        'risk_bins': ['140以上160未満', '160以上180未満', '180以上'],
        'unit': 'mg/dL',
    },
    '尿蛋白': {
        'risk_key': 'urine_protein_ge_1plus',
        'risk_label': '尿蛋白 1+以上 (CKDリスク)',
        'risk_bins': ['＋', '＋＋', '＋＋＋'],
        'unit': '定性',
    },
}


def normalize_bin(s):
    if s is None: return None
    return str(s).strip().replace('\u3000', '').replace('\xa0', '')


def num(v):
    return int(v) if isinstance(v, (int, float)) else 0


def extract_bins_v2(xlsx_path, metric):
    """Excel から (pref, bin_label, sex, age_group, count) を全展開"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = str(pref_cell).strip()
        if not last_pref: continue
        bin_lbl = normalize_bin(ws.cell(r, 2).value)
        if not bin_lbl: continue
        # 男 7階級 (col 3-9)
        for i, age in enumerate(AGE_GROUPS):
            v = num(ws.cell(r, 3 + i).value)
            if v > 0:
                records.append({
                    'pref': last_pref, 'metric': metric,
                    'bin_label': bin_lbl, 'sex': 'male',
                    'age_group': age, 'count': v,
                })
        # 女 7階級 (col 11-17)
        for i, age in enumerate(AGE_GROUPS):
            v = num(ws.cell(r, 11 + i).value)
            if v > 0:
                records.append({
                    'pref': last_pref, 'metric': metric,
                    'bin_label': bin_lbl, 'sex': 'female',
                    'age_group': age, 'count': v,
                })
    wb.close()
    return records


def build_standard_pop(all_records):
    """NDB内標準人口: 全国合算 sex × age_group の denominator 構成比
    metricごとに微差はあるが、BMI を主標準として使用
    (BMI は全受診者でほぼ全数が記録されているため代表値として適切)
    """
    bmi_recs = [r for r in all_records if r['metric'] == 'BMI']
    
    # 全国 (47県+判別不可) sex × age_group の合計
    pop_by_sex_age = collections.defaultdict(int)
    for r in bmi_recs:
        if r['pref'] == '都道府県判別不可': continue  # 標準人口は47県のみ
        key = (r['sex'], r['age_group'])
        pop_by_sex_age[key] += r['count']
    
    total = sum(pop_by_sex_age.values())
    
    std_pop = {
        'source': 'NDB第10回 特定健診 BMI集計から導出 (全47県合算 sex × age_group)',
        'method': '47県の40-74歳健診受診者を標準集団として、性年齢階級の構成比を算出',
        'total_population': total,
        'distribution': {},
    }
    for (sex, age), pop in sorted(pop_by_sex_age.items()):
        std_pop['distribution'][f'{sex}_{age}'] = {
            'sex': sex,
            'age_group': age,
            'population': pop,
            'weight': round(pop / total * 1e6) / 1e6,
        }
    
    # 構成比合計検証
    weight_sum = sum(v['weight'] for v in std_pop['distribution'].values())
    std_pop['weight_sum_check'] = round(weight_sum, 6)
    return std_pop


def compute_standardized_rates(all_records, std_pop):
    """各県・各metricのリスク率を粗率と年齢標準化率で算出"""
    std_dist = std_pop['distribution']
    
    # 県別 metric × sex × age_group × bin_label への集約
    pref_metric_sex_age = collections.defaultdict(int)
    for r in all_records:
        if r['pref'] == '都道府県判別不可': continue
        key = (r['pref'], r['metric'], r['sex'], r['age_group'], r['bin_label'])
        pref_metric_sex_age[key] += r['count']
    
    # 県別 metric × sex × age_group の denominator (全bin合計)
    pref_metric_sex_age_total = collections.defaultdict(int)
    for (pref, metric, sex, age, bl), cnt in pref_metric_sex_age.items():
        pref_metric_sex_age_total[(pref, metric, sex, age)] += cnt
    
    # リスク率算出
    results = {}
    for metric, rd in RISK_DEFS.items():
        risk_bins = set(rd['risk_bins'])
        rk = rd['risk_key']
        results[rk] = {
            'metric': metric,
            'risk_label': rd['risk_label'],
            'risk_bins': rd['risk_bins'],
            'unit': rd['unit'],
            'standardization_method': 'NDB internal standard population (47県合算 sex × age_group)',
            'by_pref': {},
        }
        
        # 全 県を集める
        prefs = sorted(set(p for (p, m, s, a) in pref_metric_sex_age_total if m == metric))
        for pref in prefs:
            # 粗率: (リスクbin の合計) / (全bin の合計)
            num_total = 0
            denom_total = 0
            # 性年齢別 stratum-specific rate を準備
            stratum_rates = {}  # key: (sex, age) → (num, denom)
            for (p, m, s, a), denom in pref_metric_sex_age_total.items():
                if p != pref or m != metric: continue
                num_in_stratum = sum(
                    cnt for (p2, m2, s2, a2, bl), cnt in pref_metric_sex_age.items()
                    if p2 == p and m2 == m and s2 == s and a2 == a and bl in risk_bins
                )
                num_total += num_in_stratum
                denom_total += denom
                stratum_rates[(s, a)] = (num_in_stratum, denom)
            
            if denom_total == 0: continue
            
            crude_rate = round(num_total / denom_total * 1000) / 10
            
            # 年齢標準化率: Σ(stratum_rate × std_weight)
            std_rate_total = 0
            std_weight_used = 0
            for (sex, age), (n, d) in stratum_rates.items():
                if d == 0: continue
                stratum_rate = n / d
                weight = std_dist.get(f'{sex}_{age}', {}).get('weight', 0)
                std_rate_total += stratum_rate * weight
                std_weight_used += weight
            
            if std_weight_used > 0:
                # weight が 1 になるように正規化 (一部 stratum 欠損対策)
                age_std_rate = round(std_rate_total / std_weight_used * 1000) / 10
            else:
                age_std_rate = None
            
            results[rk]['by_pref'][pref] = {
                'crude_rate': crude_rate,
                'age_standardized_rate': age_std_rate,
                'delta_pp': round((age_std_rate - crude_rate) * 10) / 10 if age_std_rate is not None else None,
                'denominator': denom_total,
                'numerator': num_total,
            }
    
    return results


def main():
    if not RAW.exists():
        print(f'ERROR: {RAW} not found.')
        return
    
    all_records = []
    for fname, metric in FILE_METRIC.items():
        xlsx = RAW / fname
        if not xlsx.exists():
            print(f'  MISSING: {fname}')
            continue
        print(f'  ETL v2: {fname} → metric={metric}')
        records = extract_bins_v2(xlsx, metric)
        all_records.extend(records)
        print(f'    records: {len(records)}')
    
    # 1. ndb_checkup_bins_v2.json
    with open(OUT_BINS_V2, 'w', encoding='utf-8') as f:
        json.dump({
            'source': 'NDB第10回オープンデータ 特定健診 検査値階層別分布 (令和4年度)',
            'source_url': 'https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000177221_00016.html',
            'schema_version': 'v2',
            'schema_diff': 'v1: 男女合算・全年齢合算 / v2: sex × age_group × bin_label を完全保持',
            'data': all_records,
        }, f, ensure_ascii=False, indent=2)
    print(f'\n保存: {OUT_BINS_V2} ({len(all_records):,} records)')
    
    # 2. ndb_checkup_standard_pop.json
    std_pop = build_standard_pop(all_records)
    with open(OUT_STD_POP, 'w', encoding='utf-8') as f:
        json.dump(std_pop, f, ensure_ascii=False, indent=2)
    print(f'保存: {OUT_STD_POP} (total={std_pop["total_population"]:,}, weight_sum={std_pop["weight_sum_check"]})')
    
    # 3. ndb_checkup_risk_rates_standardized.json
    std_rates = compute_standardized_rates(all_records, std_pop)
    with open(OUT_RATES_STD, 'w', encoding='utf-8') as f:
        json.dump({
            'source': 'NDB第10回オープンデータ 特定健診 リスク該当者率 (粗率 + 年齢標準化率)',
            'method': 'direct standardization with NDB internal standard population (47県合算 sex × age_group)',
            'risk_rates': std_rates,
        }, f, ensure_ascii=False, indent=2)
    print(f'保存: {OUT_RATES_STD} ({len(std_rates)} risk metrics)')
    
    # サマリ + 5県 sanity
    print('\n=== Phase 2C-1 標準化サマリ ===')
    print(f'\n{"県":<8}{"metric":<12}{"粗率":>10}{"標準化率":>12}{"delta":>10}')
    target_prefs = ['東京都','大阪府','北海道','沖縄県','高知県']
    for pref in target_prefs:
        for rk in ['bmi_ge_25', 'hba1c_ge_6_5', 'sbp_ge_140', 'ldl_ge_140', 'urine_protein_ge_1plus']:
            entry = std_rates[rk]['by_pref'].get(pref)
            if entry:
                metric = std_rates[rk]['metric']
                cr = entry['crude_rate']
                ar = entry['age_standardized_rate']
                d = entry['delta_pp']
                print(f'{pref:<8}{metric:<12}{cr:>9.1f}%{ar:>11.1f}%{d:>+9.1f}pp')
        print('-'*58)


if __name__ == '__main__':
    main()
