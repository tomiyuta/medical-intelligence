#!/usr/bin/env python3
"""
phase3 ETL: NDB特定健診質問票 Q18/Q19/Q20 を ndb_questionnaire.json に追加

データ出典: 厚労省 第10回NDBオープンデータ 特定健診 質問票（令和4年度）
- 815.xlsx: Q18 飲酒頻度（毎日/時々/ほとんど飲まない）
- 817.xlsx: Q19 飲酒量（1合未満/1〜2合未満/2〜3合未満/3合以上） ※飲酒者のみ回答
- 820.xlsx: Q20 睡眠で休養が十分とれている（はい/いいえ）

派生指標:
- drinking_daily: Q18 「毎日」/ Q18合計 = 毎日飲酒率
- heavy_drinker:  Q19 「2合以上」(2〜3合未満+3合以上) / Q19合計 = 高量飲酒率（飲酒者中）
- sleep_ok:       Q20 「はい」/ Q20合計 = 睡眠充足率（高=低リスク・xInverse対象）

列レイアウト（共通）:
  A: 都道府県名 / B: 回答 / C-I: 男40-44..70-74 / J: 男中計 / K-Q: 女40-44..70-74 / R: 女中計
  ヘッダ5行 / データはR6から
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / 'data' / 'raw_ndb_questionnaire'
OUT = ROOT / 'data' / 'static' / 'ndb_questionnaire.json'

COL_MALE_SUBTOTAL = 10
COL_FEMALE_SUBTOTAL = 18


def read_question(filename):
    wb = openpyxl.load_workbook(RAW / filename, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result = {}
    last_pref = None
    for r in range(6, ws.max_row + 1):
        pref_cell = ws.cell(r, 1).value
        if pref_cell:
            last_pref = pref_cell
        if not last_pref:
            continue
        answer = ws.cell(r, 2).value
        if not answer:
            continue
        male = ws.cell(r, COL_MALE_SUBTOTAL).value
        female = ws.cell(r, COL_FEMALE_SUBTOTAL).value
        m = male if isinstance(male, (int, float)) else 0
        f = female if isinstance(female, (int, float)) else 0
        result.setdefault(last_pref, {})[answer] = m + f
    wb.close()
    return result


def compute_rate(dist, target_answers):
    total = sum(dist.values())
    if total <= 0:
        return None
    target_sum = sum(dist.get(a, 0) for a in target_answers)
    return round(target_sum / total * 1000) / 10


def main():
    print('Reading existing JSON...')
    with open(OUT, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f'  prefs: {len(data["prefectures"])}, questions: {list(data["questions"].keys())}')

    print('Loading Q18 (815.xlsx) — 飲酒頻度...')
    q18 = read_question('815.xlsx')
    print('Loading Q19 (817.xlsx) — 飲酒量...')
    q19 = read_question('817.xlsx')
    print('Loading Q20 (820.xlsx) — 睡眠...')
    q20 = read_question('820.xlsx')

    print(f'Q18 prefs: {len(q18)} | answers: {sorted(set().union(*[d.keys() for d in q18.values()]))}')
    print(f'Q19 prefs: {len(q19)} | answers: {sorted(set().union(*[d.keys() for d in q19.values()]))}')
    print(f'Q20 prefs: {len(q20)} | answers: {sorted(set().union(*[d.keys() for d in q20.values()]))}')

    Q18_DAILY = ['毎日']
    Q19_HEAVY = ['2～3合未満', '3合以上']
    Q20_OK = ['はい']

    updated = 0
    for pref in data['prefectures'].keys():
        d = data['prefectures'][pref]
        if pref in q18:
            r = compute_rate(q18[pref], Q18_DAILY)
            if r is not None:
                d['drinking_daily'] = r
        if pref in q19:
            r = compute_rate(q19[pref], Q19_HEAVY)
            if r is not None:
                d['heavy_drinker'] = r
        if pref in q20:
            r = compute_rate(q20[pref], Q20_OK)
            if r is not None:
                d['sleep_ok'] = r
        updated += 1

    data['questions']['drinking_daily'] = {
        'label': '毎日飲酒',
        'question': 'お酒を毎日飲む（特定健診Q18）',
        'risk_label': '毎日飲酒率',
    }
    data['questions']['heavy_drinker'] = {
        'label': '高量飲酒',
        'question': '飲酒日1日2合以上（特定健診Q19・飲酒者中）',
        'risk_label': '2合以上飲酒率',
    }
    data['questions']['sleep_ok'] = {
        'label': '睡眠充足',
        'question': '睡眠で休養が十分とれている（特定健診Q20）',
        'risk_label': '睡眠充足率',
    }

    print(f'Updated {updated} prefs.')
    samples = ['北海道', '東京都', '大阪府', '沖縄県', '京都府']
    for p in samples:
        if p in data['prefectures']:
            v = data['prefectures'][p]
            print(f'  {p}: daily={v.get("drinking_daily")}% heavy={v.get("heavy_drinker")}% sleep={v.get("sleep_ok")}%')

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
