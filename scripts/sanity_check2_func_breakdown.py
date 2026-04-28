#!/usr/bin/env python3
"""R6 様式1 の機能区分カラムを特定し、機能別床数集計の妥当性を確認"""
import openpyxl
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
SAMPLE = ROOT / 'data' / 'raw' / 'source' / '06_病床機能報告' / 'data_R6' / 'R6_様式1_北海道東北.xlsx'

print('Loading 北海道東北 様式1...')
wb = openpyxl.load_workbook(SAMPLE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Get all rows for inspection
all_rows = list(ws.iter_rows(values_only=True))
print(f'Total rows: {len(all_rows)}')

# Find 機能区分 cell content for col 15, 16 (and search broadly)
# Sample data rows (R6+) — see what's in col 15, 16
print('\n=== Col 15-16 sample values (first 10 data rows) ===')
for r_idx in range(5, min(15, len(all_rows))):
    row = all_rows[r_idx]
    print(f'  R{r_idx+1}: col15="{row[15]}" col16="{row[16]}"')

# 集計: col15(2024時点) で 機能区分 → 一般+療養床数 合計
agg = defaultdict(lambda: {'wards': 0, 'beds': 0})
for r_idx in range(5, len(all_rows)):
    row = all_rows[r_idx]
    if not row or len(row) < 23:
        continue
    func = row[15]
    if not func:
        continue
    func_str = str(func).strip()
    ipan = row[18] if isinstance(row[18], (int, float)) else 0
    ryoyo = row[22] if isinstance(row[22], (int, float)) else 0
    agg[func_str]['wards'] += 1
    agg[func_str]['beds'] += ipan + ryoyo

print('\n=== 北海道+東北 機能区分別集計（col 15: 2024年7月1日時点） ===')
total_b = 0
total_w = 0
for func, v in sorted(agg.items(), key=lambda x: -x[1]['beds']):
    print(f'  {func[:30]:<30}: wards={v["wards"]:>5} beds={v["beds"]:>8,}')
    total_b += v['beds']
    total_w += v['wards']
print(f'  ---')
print(f'  合計: wards={total_w} beds={total_b:,}')

wb.close()
